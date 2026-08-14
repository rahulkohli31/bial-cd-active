"""Turn BIAL's vulnerability exports into a remediation report the reviewer can add up.

THE GENERATOR IS TRACKED; ITS INPUTS AND OUTPUTS ARE NOT. Both the incoming vulnerability
lists and the workbook this produces are client vulnerability data about a live system, and
this repository is public. They live in a local working directory OUTSIDE the repo tree; the
repo's ignore rules are a backstop against an accidental `git add`, not the control. Only this
file — code, which has to be reviewable and testable — belongs in the tree.

That split is also why THERE ARE NO CVE IDENTIFIERS IN THIS FILE. Every disposition rule below
matches on STRUCTURE — the image, the package manager, the software name, the installation path
— never on a list of specific findings. A CVE list would be client data, would rot the moment
BIAL rescans, and would silently stop matching without failing. A structural rule keeps working
across rescans and states the actual reason a finding is cleared, which is what a reviewer is
checking.

TWO PASSES, ONE TOOL.

  Pass 1 (the coverage map — plan U9). Run against BIAL's BEFORE export alone, before any
  Dockerfile is touched. Every row is mapped to its INTENDED disposition: cleared by a named
  unit, exception, dispute, deferred, or held. That map is the checkable definition of done —
  at handover no row may be unmapped — and it is what bounds the remediation work. A fix made
  before the map exists cannot be attributed to a row, and unattributed fixes are exactly what
  make a reduction claim unverifiable.

      uv run python -m scripts.exception_register map \\
          --before "<dir>/citizen-dev-sandbox - vulnerabilities.xlsx" \\
          --before "<dir>/vibe-coding_sheet.xlsx" \\
          --out-dir "<dir>/coverage"

  Pass 2 (the register — plan U7). Run again once BIAL returns the post-remediation export.
  Every BEFORE row is reconciled against what actually survived: rows that disappeared are
  Fixed, rows that survived carry their residual disposition, and a residual row the map never
  anticipated is reported as a NON-RECONCILING row — a signal the fix did not land, not a new
  exception.

      uv run python -m scripts.exception_register register \\
          --before ... --after "<dir>/post-remediation.xlsx" \\
          --out "<dir>/out/bial-remediation-report.xlsx"

THE PARTITION INVARIANT IS THE POINT. Every row on the original list lands in exactly one
sheet. The sheets' accounted-row counts sum back to the original total, and the residual sheets
sum to the post-remediation total. A reviewer who adds the sheets up and gets a different number
has found a real error, so the arithmetic is the report's own self-check rather than decoration.
`Summary` is therefore never typed: every number on it is an Excel formula over the other
sheets, so the report cannot drift internally even if someone edits a row by hand.

GRANULARITY (plan AE6). One ENTRY per (image, CVE, software name, software version), each
carrying `Rows Accounted` — the number of scanner rows that collapse into it. BIAL's scanner
emits one row per affected sibling package AND per installation path, so 1,083 rows resolve to
far fewer real findings; reporting entries while carrying the row count lets the reviewer
reconcile against the row totals their own console shows, in either direction.
"""

from __future__ import annotations

import argparse
import csv
import functools
import json
import re
import sys
from collections import Counter, defaultdict
from collections.abc import Iterable, Sequence
from dataclasses import dataclass, field
from datetime import UTC, datetime
from enum import StrEnum
from pathlib import Path
from typing import Any, Final, TextIO

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# ─────────────────────────────────────────────────────────────────────────────
# Dispositions
# ─────────────────────────────────────────────────────────────────────────────


class Disposition(StrEnum):
    """What we are saying about a finding. Exactly one applies to every row.

    The five RESIDUAL states are the ones the handover defends: EXCEPTION, DISPUTE,
    DEFERRED, SUPERSEDED and DISPUTE_REJECTED. FIXED is not residual (it is gone from the
    after-export) and OUT_OF_SCOPE is attributed elsewhere rather than defended.

    HELD is a PRE-FINAL state and must be empty before delivery — it exists so that rows we
    cannot honestly disposition yet are visible and counted rather than quietly filed as
    something they are not. A held row makes both CLI commands exit non-zero, so a report that
    is not ready to ship cannot be mistaken for one that is.
    """

    FIXED = "fixed"
    EXCEPTION = "exception"
    DEFERRED = "deferred"
    DISPUTE = "dispute"
    DISPUTE_REJECTED = "dispute-rejected"
    SUPERSEDED = "superseded-artifact"
    OUT_OF_SCOPE = "out-of-scope"
    HELD = "held-pending-answer"


#: Sheet each disposition is written to. Two dispositions may share a sheet — DISPUTE_REJECTED
#: falls back to accepted risk and so belongs beside the exceptions it joins, and a superseded
#: artifact is attributed away from the shipped image exactly like an out-of-scope one. The
#: partition invariant is over ROWS, not over dispositions, and every sheet carries the precise
#: `Disposition` column so nothing is blurred by the shared home.
SHEET_FOR: Final[dict[Disposition, str]] = {
    Disposition.FIXED: "Fixed",
    Disposition.EXCEPTION: "Exceptions",
    Disposition.DISPUTE_REJECTED: "Exceptions",
    Disposition.DEFERRED: "Deferred",
    Disposition.DISPUTE: "Disputes",
    Disposition.SUPERSEDED: "Out of scope",
    Disposition.OUT_OF_SCOPE: "Out of scope",
    Disposition.HELD: "Held",
}

#: Sheet order in the workbook. `Summary` is written first so it is what opens.
CONTENT_SHEETS: Final[tuple[str, ...]] = (
    "Fixed",
    "Exceptions",
    "Deferred",
    "Disputes",
    "Out of scope",
    "Held",
)

#: NOTE: the plan's second reconciliation direction — "the residual sheets sum back to the
#: post-remediation export's total" — is NOT implemented, and a `RESIDUAL_SHEETS` constant that
#: named the sheets without checking anything only made it look like it was. It cannot be
#: implemented against the current entry model: a surviving entry carries its BEFORE rows, so
#: summing the residual sheets yields the before-count of the residual set, not the after-count.
#: Doing it properly needs each surviving entry to carry its after-row count alongside its
#: before-row count. Tracked as follow-up rather than faked here.


# ─────────────────────────────────────────────────────────────────────────────
# Version comparison
# ─────────────────────────────────────────────────────────────────────────────

_EPOCH = re.compile(r"^(\d+):")
#: Pre-release markers, anchored so an Alpine revision (`-r5`) and a Debian point release
#: (`+deb12u15`) are never mistaken for one. `~` is Debian's own pre-release marker and sorts
#: before everything, including the empty string.
_PRERELEASE = re.compile(
    r"(?:(?<=\d)(?:a|b|rc)\d)|(?:[-._](?:alpha|beta|rc|dev|pre)[-._\d]?)|~",
    re.IGNORECASE,
)


def is_prerelease(version: str) -> bool:
    """True when `version` is an alpha/beta/rc/dev build rather than a released one.

    Load-bearing for plan AE3: a finding whose ONLY offered fix is a pre-release cannot be
    taken (Scope Boundaries forbids pre-release language runtimes), so it becomes accepted
    risk WITH the pre-release status as its stated reason — never filed as not-affected.
    """
    return bool(_PRERELEASE.search(version.strip()))


def _split_debian(version: str) -> tuple[int, str, str]:
    """Split into (epoch, upstream, revision) the way dpkg does."""
    v = version.strip()
    epoch = 0
    if m := _EPOCH.match(v):
        epoch = int(m.group(1))
        v = v[m.end() :]
    upstream, sep, revision = v.rpartition("-")
    if not sep:  # no revision at all
        upstream, revision = v, ""
    return epoch, upstream, revision


def _order(ch: str) -> int:
    """dpkg's character collation: `~` < end-of-string < letters < everything else."""
    if ch == "~":
        return -1
    if ch.isdigit():
        return 0  # digits are never compared here; the caller splits them out first
    if ch.isalpha():
        return ord(ch)
    return ord(ch) + 256


def _compare_part(a: str, b: str) -> int:
    """Compare one dpkg version part (upstream or revision). Returns -1, 0 or 1."""
    i = j = 0
    while i < len(a) or j < len(b):
        # Non-digit run, collated by dpkg's modified ordering.
        first_diff = 0
        while (i < len(a) and not a[i].isdigit()) or (j < len(b) and not b[j].isdigit()):
            ac = _order(a[i]) if i < len(a) and not a[i].isdigit() else 0
            bc = _order(b[j]) if j < len(b) and not b[j].isdigit() else 0
            if ac != bc:
                first_diff = -1 if ac < bc else 1
                break
            if i < len(a) and not a[i].isdigit():
                i += 1
            if j < len(b) and not b[j].isdigit():
                j += 1
        if first_diff:
            return first_diff
        # Numeric run, compared as integers so 10 > 9.
        start_i, start_j = i, j
        while i < len(a) and a[i].isdigit():
            i += 1
        while j < len(b) and b[j].isdigit():
            j += 1
        na = int(a[start_i:i] or 0)
        nb = int(b[start_j:j] or 0)
        if na != nb:
            return -1 if na < nb else 1
        if (start_i, start_j) == (i, j):  # neither side advanced — done
            break
    return 0


def compare_versions(a: str, b: str) -> int:
    """Compare two package versions using dpkg's algorithm. Returns -1, 0 or 1.

    dpkg's rules are the right general tool here even for Alpine and language-ecosystem
    versions: they handle an epoch, a `-rN` revision, a `+deb12u15` point release and a `~rc1`
    pre-release without special-casing, and `packaging.version` rejects most of those outright.
    """
    ea, ua, ra = _split_debian(_strip_prefix(a))
    eb, ub, rb = _split_debian(_strip_prefix(b))
    if ea != eb:
        return -1 if ea < eb else 1
    if (c := _compare_part(ua, ub)) != 0:
        return c
    return _compare_part(ra, rb)


def _strip_prefix(version: str) -> str:
    """Drop a leading `go` or `v` so `go1.22.3` compares against a bare `1.22.4`."""
    v = version.strip()
    for prefix in ("go", "v", "V"):
        if v.startswith(prefix) and v[len(prefix) : len(prefix) + 1].isdigit():
            return v[len(prefix) :]
    return v


def parse_fix_versions(fixed: str) -> list[str]:
    """Split a `Fixed Version` cell into candidate versions.

    The Go feed offers one fix per still-supported release line —
    `1.24.13, 1.25.7, 1.26.0-rc.3` — while deb/apk offer exactly one. `-` means the feed has
    no fix at all, which is not the same as "not affected" and must never be reported as such.
    """
    raw = (fixed or "").strip()
    if raw in {"", "-", "None", "n/a", "N/A"}:
        return []
    return [p.strip() for p in re.split(r"[,;]| or ", raw) if p.strip()]


def _line_of(version: str) -> str:
    """The `major.minor` release line a version sits on."""
    parts = _strip_prefix(version).split(".")
    return ".".join(parts[:2]) if len(parts) >= 2 else _strip_prefix(version)


def already_at_or_past_fix(installed: str, fixed: str) -> bool:
    """True when the INSTALLED version already meets or exceeds the vendor's fixed version.

    This is plan AE2, and it routes to DISPUTE rather than to an exception: if we already ship
    the fixed version, the finding is a scanner error, and filing it as an accepted risk would
    concede a vulnerability we do not have.

    The release-LINE rule matters and is easy to get wrong. Installed `go1.23.12` against fixes
    `1.24.9, 1.25.3` is NOT disputed — the vendor is saying "there is no fix on your line, move
    up" — whereas installed `go1.22.3` against `1.21.11, 1.22.4` compares only against the
    1.22 candidate. Comparing against the maximum candidate in either case would mark almost
    every genuine finding as a dispute, which is the single most damaging error this report can
    make: it is the one claim the reviewer can disprove from their own console.
    """
    candidates = parse_fix_versions(fixed)
    if not candidates or not installed.strip():
        return False
    line = _line_of(installed)
    same_line = [c for c in candidates if _line_of(c) == line]
    # Same-line fixes decide it. With none, the vendor offers nothing on our line, so we are
    # only "past" the fix if we are past ALL of them (a genuinely newer major).
    pool = same_line or candidates
    target = max(pool, key=lambda c: _VersionKey(c))
    return compare_versions(installed, target) >= 0


@functools.total_ordering
@dataclass(frozen=True, slots=True)
class _VersionKey:
    """`max()` helper — dpkg ordering is a comparison, not a sortable tuple.

    `total_ordering` derives the rest from `__lt__`. With only `__lt__` defined this still
    worked, but by an implicit route: `max()` evaluates `item > best`, finds no `__gt__`, and
    silently retries as `best.__lt__(item)`. Correct, and one `__gt__` away from breaking in a
    way no test would obviously localise.
    """

    version: str

    def __lt__(self, other: _VersionKey) -> bool:
        return compare_versions(self.version, other.version) < 0


# ─────────────────────────────────────────────────────────────────────────────
# The normalised finding
# ─────────────────────────────────────────────────────────────────────────────


@dataclass(frozen=True, slots=True)
class Finding:
    """One scanner ROW, normalised across both of BIAL's export shapes.

    The two workbooks have different column shapes — the sandbox export carries installation
    path, EPSS and software type; the other does not — so the UNION is the working schema, with
    the missing columns left empty rather than dropped. Dropping them would silently discard
    the installation path, which is the only thing that separates the sandbox's Go-toolchain
    findings (vendored build binaries) from its proxy findings and its OS findings.
    """

    image: str
    asset: str
    cve_id: str
    severity: str
    software_name: str
    software_version: str
    fixed_version: str
    package_manager: str
    software_type: str
    install_path: str
    epss: str
    exploited: str
    first_detected: str
    status: str
    description: str
    source: str

    @property
    def key(self) -> tuple[str, str, str, str]:
        """The ENTRY key (plan AE6): one entry per image, CVE, package name and version."""
        return (self.image, self.cve_id, self.software_name, self.software_version)

    @property
    def fixable(self) -> bool:
        """True when the vendor feed names at least one RELEASED fixed version.

        A fix that exists only as a pre-release is deliberately NOT fixable here: Scope
        Boundaries rules out pre-release language runtimes, so counting it as fixable would
        make the zero-fixable commitment false on its own terms.
        """
        candidates = parse_fix_versions(self.fixed_version)
        return any(not is_prerelease(c) for c in candidates)

    @property
    def prerelease_only_fix(self) -> bool:
        """True when a fix exists but EVERY offered fix is a pre-release (plan AE3)."""
        candidates = parse_fix_versions(self.fixed_version)
        return bool(candidates) and all(is_prerelease(c) for c in candidates)


# ─────────────────────────────────────────────────────────────────────────────
# Ingest
# ─────────────────────────────────────────────────────────────────────────────

#: Header names that identify the two shapes. Detection is by HEADER, never by filename — a
#: file renamed on the way out of an inbox must not change how it parses.
_ASSET_SHEET_MARKERS: Final[frozenset[str]] = frozenset({"Asset Name", "Severity"})
_SEVERITY_SHEET_NAMES: Final[tuple[str, ...]] = ("Critical", "High", "Medium", "Low")

_ASSET_RE = re.compile(r"^(?P<registry>[^/]+)/(?P<repo>[^@:]+)(?:@(?P<digest>sha256:[0-9a-f]+))?")


def image_key(asset: str) -> str:
    """The short repository name used as the image identity throughout the report."""
    if m := _ASSET_RE.match(str(asset or "").strip()):
        return m.group("repo")
    return str(asset or "").strip() or "unknown"


def asset_digest(asset: str) -> str:
    """The manifest digest embedded in a scanner asset reference, if it carries one.

    BIAL's export names each scanned asset by DIGEST, not by tag. That is the rollback anchor
    for every image the scan covered, recoverable without any registry access at all — which
    matters, because all four artifacts ship under mutable tags and the first push untags the
    manifest the previous release ran.
    """
    if m := _ASSET_RE.match(str(asset or "").strip()):
        return m.group("digest") or ""
    return ""


def _cell(value: Any) -> str:
    """Normalise a cell to a trimmed string; `None` and the feed's `-` become empty."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    return str(value).strip()


def _header_row(ws: Worksheet) -> tuple[int, list[str]]:
    """Find the header row and return (row_index, headers).

    BIAL's pivot-style `Summary` sheets carry blank leading rows, so the first row is not
    reliably the header; scan the first few rows for one that looks like a finding header.
    """
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=6, values_only=True), start=1):
        headers = [_cell(c) for c in row]
        if "CVE ID" in headers:
            return idx, headers
    return 0, []


def _sheet_asset(wb: Workbook) -> str:
    """Pull the asset reference out of a workbook whose finding sheets do not carry one.

    The sandbox export splits by SEVERITY, so the asset — and with it the digest — appears
    only on the pivot `Summary` sheet. Losing it would leave every sandbox row unattributed.
    """
    if "Summary" not in wb.sheetnames:
        return ""
    for row in wb["Summary"].iter_rows(min_row=1, max_row=40, values_only=True):
        for cell in row:
            text = _cell(cell)
            if "/" in text and ("azurecr.io" in text or text.startswith("sha256:")):
                return text
    return ""


def load_export(path: Path) -> list[Finding]:
    """Normalise one BIAL export workbook into scanner rows.

    Handles both shapes: severity-per-sheet (asset on the pivot Summary) and asset-per-sheet
    (severity as a column). Pivot/summary sheets are skipped — they are a view of the rows,
    and counting them would double the totals.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        fallback_asset = _sheet_asset(wb)
        out: list[Finding] = []
        for ws in wb.worksheets:
            hdr_idx, headers = _header_row(ws)
            if not headers:
                continue  # a pivot/summary sheet, not findings
            col = {name: i for i, name in enumerate(headers) if name}
            sheet_severity = ws.title if ws.title in _SEVERITY_SHEET_NAMES else ""
            for row_idx, row in enumerate(
                ws.iter_rows(min_row=hdr_idx + 1, values_only=True), start=hdr_idx + 1
            ):
                cells = [_cell(c) for c in row]
                if not any(cells):
                    continue

                def get(name: str, _cells: list[str] = cells) -> str:
                    i = col.get(name)
                    return _cells[i] if i is not None and i < len(_cells) else ""

                cve = get("CVE ID")
                if not cve:
                    continue
                asset = get("Asset Name") or fallback_asset
                out.append(
                    Finding(
                        image=image_key(asset),
                        asset=asset,
                        cve_id=cve,
                        severity=get("Severity") or sheet_severity,
                        software_name=get("Software Name"),
                        software_version=get("Software Version"),
                        fixed_version=get("Fixed Version"),
                        package_manager=get("Software Package Manager"),
                        software_type=get("Software Type"),
                        install_path=get("Installation Path"),
                        epss=get("CVE EPSS Score"),
                        exploited=get("Exploited in the Wild"),
                        first_detected=get("First Detected"),
                        status=get("Status"),
                        description=get("Description"),
                        source=f"{path.name}#{ws.title}#{row_idx}",
                    )
                )
        return out
    finally:
        wb.close()


def load_exports(paths: Sequence[Path]) -> list[Finding]:
    """Normalise several exports into one table."""
    findings: list[Finding] = []
    for p in paths:
        findings.extend(load_export(p))
    return findings


# ─────────────────────────────────────────────────────────────────────────────
# Disposition rules — STRUCTURAL, never a CVE list
# ─────────────────────────────────────────────────────────────────────────────


@dataclass(frozen=True, slots=True)
class Rule:
    """One structural matcher. The first rule that matches a finding wins."""

    name: str
    disposition: Disposition
    reason: str
    images: tuple[str, ...] = ()
    unit: str = ""
    owner: str = ""
    target_date: str = ""
    package_managers: tuple[str, ...] = ()
    software_names: tuple[str, ...] = ()
    path_contains: tuple[str, ...] = ()
    path_excludes: tuple[str, ...] = ()
    #: None = don't care · True = only when a RELEASED fix exists · False = only when none does.
    requires_fix: bool | None = None

    def matches(self, f: Finding) -> bool:
        if self.images and not any(f.image.endswith(i) or i in f.image for i in self.images):
            return False
        if self.package_managers and f.package_manager not in self.package_managers:
            return False
        if self.software_names and f.software_name not in self.software_names:
            return False
        if self.path_contains and not any(p in f.install_path for p in self.path_contains):
            return False
        if self.path_excludes and any(p in f.install_path for p in self.path_excludes):
            return False
        if self.requires_fix is not None and f.fixable is not self.requires_fix:
            return False
        return True


SANDBOX: Final = "citizen-dev-sandbox"
FRONTEND: Final = "vibe-coding-frontend"
BACKEND: Final = "vibe-coding-backend"

#: Image libraries that `nginx:alpine-slim` does not ship at all. Removing a package is the
#: strongest possible remediation — the finding cannot come back at any version — so these are
#: called out separately from the base-version bump they arrive with.
_IMAGE_LIBS: Final[tuple[str, ...]] = (
    "tiff",
    "libtiff",
    "libpng",
    "libpng16",
    "libxpm",
    "freetype",
    "libjpeg",
    "libjpeg-turbo",
    "libwebp",
    "libavif",
    "gd",
    "libgd",
)

#: Retained deliberately in the sandbox (plan Key Technical Decisions). The command surface
#: there is an LLM with unrestricted command execution, and the supervisor's own kill denylist
#: is evidence the agent reaches for process tools. Only the rows with NO released fix become
#: accepted risk — a fixable one is still cleared by the base move, and filing THAT as an
#: exception would be the falsifiable claim the reviewer catches.
_SANDBOX_KEPT_TOOLS: Final[tuple[str, ...]] = (
    "curl",
    "libcurl4",
    "libcurl",
    "procps",
    "libprocps8",
    "libproc2-0",
)

RULES: Final[tuple[Rule, ...]] = (
    # ── sandbox: vendored Go build binaries ──────────────────────────────────
    Rule(
        name="sandbox-caddy",
        images=(SANDBOX,),
        path_contains=("/usr/local/bin/caddy",),
        disposition=Disposition.FIXED,
        unit="U3",
        reason="Caddy static binary pinned forward (v2.8.4 → v2.11.4), rebuilding its Go "
        "toolchain and vendored modules. This is the only genuinely network-facing "
        "component in the image.",
    ),
    Rule(
        name="sandbox-esbuild-kit-dead-chain",
        images=(SANDBOX,),
        path_contains=("@esbuild-kit",),
        disposition=Disposition.FIXED,
        unit="U3",
        reason="The nested esbuild 0.18.20 (Go 1.20) copies under the deprecated "
        "@esbuild-kit loader chain are collapsed onto the overridden version. The "
        "@esbuild-kit packages themselves remain in the tree — it is the vendored Go "
        "binaries they carried, which were scanned but never executed, that are gone.",
    ),
    Rule(
        name="sandbox-esbuild-live",
        images=(SANDBOX,),
        path_contains=("/node_modules/esbuild", "/node_modules/@esbuild/"),
        disposition=Disposition.FIXED,
        unit="U3",
        reason="esbuild collapsed onto a single current version by a package override, "
        "rebuilding the vendored Go binary that runs schema generation.",
    ),
    # ── sandbox: the package manager's OWN bundled modules (plan R3) ─────────
    Rule(
        name="sandbox-bundled-npm",
        images=(SANDBOX,),
        path_contains=("/usr/local/lib/node_modules/npm/",),
        disposition=Disposition.FIXED,
        unit="U3",
        reason="Vendor-bundled application packages inside npm itself, upgraded with the "
        "bundled package manager in the image (R3). These are not project dependencies "
        "and no lockfile change reaches them.",
    ),
    #: CONSTRAINED TO THE PATH ITS REASON NAMES. Without `path_contains` this matched EVERY
    #: python-package-manager row on the sandbox image — including the supervisor's own
    #: pip-installed FastAPI/uvicorn/starlette under site-packages, which the Debian base move
    #: does not touch at all. Filing those as cleared-by-the-base would be a claim the reviewer
    #: can disprove from their own console, and it is the precise shape of error this file's
    #: docstring warns about: a rule broader than the reason attached to it.
    #:
    #: The supervisor's pinned dependencies now fall through to the fixability fallback, which
    #: is the honest answer — a fixable one lands in DEFERRED with UNASSIGNED/UNSET and is loud
    #: until somebody owns it, rather than quietly counted as remediated.
    Rule(
        name="sandbox-bundled-python-build-tooling",
        images=(SANDBOX,),
        package_managers=("python",),
        path_contains=("/usr/lib/python3/dist-packages",),
        disposition=Disposition.FIXED,
        unit="U3",
        reason="Distribution-bundled Python build tooling under /usr/lib/python3/"
        "dist-packages, carried forward by the Debian 13 base (R3).",
    ),
    # ── sandbox: the golden template ─────────────────────────────────────────
    #: NAMES THE PACKAGES THAT ACTUALLY MOVED, and nothing else. Plan U3 scoped a 25-of-26 pin
    #: bump to latest stable; what shipped is `next` 16.2.10 → 16.2.12 plus the `overrides`
    #: block (esbuild — its own rule above — and postcss). A `path_contains` rule over the whole
    #: of /workspace/app/node_modules/ would have filed all 26 pins as FIXED, claiming 25
    #: upgrades that were never made. Every other template dependency falls through to the
    #: fixability fallback: a fixable one becomes an owned DEFERRED, which is exactly what an
    #: un-taken pin bump is.
    #:
    #: WHEN THE REST OF U3 LANDS, widen `software_names` in the same commit as the package.json
    #: change — never ahead of it.
    Rule(
        name="sandbox-golden-template",
        images=(SANDBOX,),
        path_contains=("/workspace/app/node_modules/",),
        software_names=("next", "postcss"),
        disposition=Disposition.FIXED,
        unit="U3",
        reason="Golden-template dependency moved forward and the lockfile regenerated in the "
        "same commit: the framework to its current patch, and postcss pinned up by a "
        "package override.",
    ),
    # ── sandbox: OS packages ─────────────────────────────────────────────────
    Rule(
        name="sandbox-os-fixable",
        images=(SANDBOX,),
        package_managers=("deb",),
        requires_fix=True,
        disposition=Disposition.FIXED,
        unit="U3",
        reason="Debian 12 (bookworm) → Debian 13 (trixie) base, taken for security-support "
        "lifecycle. Bookworm left regular security support on 12 July 2026.",
    ),
    Rule(
        name="sandbox-kept-tooling-no-fix",
        images=(SANDBOX,),
        software_names=_SANDBOX_KEPT_TOOLS,
        requires_fix=False,
        disposition=Disposition.EXCEPTION,
        unit="U3",
        reason="Retained deliberately. The sandbox command surface is an LLM with "
        "unrestricted command execution, and removing these tools breaks the agent's "
        "own workflow rather than the attacker's. No upstream fix at the pinned base.",
    ),
    #: Everything left on the sandbox's Debian feed with no named fix. HELD rather than
    #: excepted, because "no fix" here is a property of the FEED, not of the package. Measured
    #: across the incoming exports: the Debian feed names a fixed version on 0 of 519 rows
    #: (0.0%), while every other feed in the same scan populates it — go-module 99.5%, npm
    #: 100%, distribution-python 100%, apk 68.2%. An empty field that is empty for every single
    #: row is not evidence of an unfixable package, so filing these as exceptions would assert
    #: something the data does not support, and it is exactly the claim a reviewer can disprove
    #: from their own console. Two things resolve them: the rescan (which shows whether the
    #: base move cleared them) and A4's answer on the feed (plan U8 question 2).
    #:
    #: The sandbox is the ONLY image where this matters. U4 removes the backend's Debian
    #: package set outright, so the feed question cannot change that image's answer.
    Rule(
        name="sandbox-os-no-fix-held",
        images=(SANDBOX,),
        package_managers=("deb",),
        requires_fix=False,
        disposition=Disposition.HELD,
        unit="U3",
        reason="No fixed version named in the Debian feed — which names one on 0 of 519 "
        "Debian rows in this scan, against 68-100% for every other feed. Held: the "
        "empty field is a property of the feed, not proof the package is unfixable. "
        "Resolved by the rescan plus A4's answer (plan U8 question 2).",
    ),
    # ── portal frontend ──────────────────────────────────────────────────────
    Rule(
        name="frontend-image-libraries-removed",
        images=(FRONTEND,),
        software_names=_IMAGE_LIBS,
        disposition=Disposition.FIXED,
        unit="U2",
        reason="Package REMOVED from the image: the -slim nginx variant does not ship the "
        "image libraries, which the portal never loads. A removed package cannot "
        "reintroduce the finding at any version.",
    ),
    Rule(
        name="frontend-nginx",
        images=(FRONTEND,),
        software_names=("nginx",),
        disposition=Disposition.FIXED,
        unit="U2",
        reason="nginx runtime moved off the minor-pinned tag family that let this base go 16 "
        "months stale, onto the current mainline Alpine slim base.",
    ),
    Rule(
        name="frontend-apk-fixable",
        images=(FRONTEND,),
        package_managers=("apk",),
        requires_fix=True,
        disposition=Disposition.FIXED,
        unit="U2",
        reason="Cleared by the Alpine base move that comes with the current nginx image.",
    ),
    Rule(
        name="frontend-apk-no-fix",
        images=(FRONTEND,),
        package_managers=("apk",),
        requires_fix=False,
        disposition=Disposition.EXCEPTION,
        unit="U2",
        reason="No fixed version in the Alpine security feed at the current release. Alpine "
        "populates that field where a fix exists, so an empty one is a real absence.",
    ),
    # ── backend ──────────────────────────────────────────────────────────────
    Rule(
        name="backend-interpreter",
        images=(BACKEND,),
        software_names=("python", "python3", "cpython"),
        disposition=Disposition.EXCEPTION,
        unit="U4",
        reason="The CPython interpreter itself. Carried at the newest RELEASED patch; the "
        "remaining findings have no released fix (see the pre-release rows, which are "
        "excluded by the plan's no-pre-release-runtimes boundary).",
    ),
    Rule(
        name="backend-debian-package-set-removed",
        images=(BACKEND,),
        package_managers=("deb",),
        disposition=Disposition.FIXED,
        unit="U4",
        reason="Package set REMOVED: the backend rebases off Debian onto Alpine, so the "
        "glibc/Debian packages carrying these findings are not present in the new "
        "image at any version. The Debian feed's fix-version question cannot change "
        "this answer, which is why these rows are not held.",
    ),
    # ── reference app (scanned only once a nominated app is measured) ────────
    Rule(
        name="reference-app-agent-dependencies",
        images=("citizen-app-",),
        path_excludes=("/usr/local/lib/node_modules/",),
        path_contains=("/node_modules/",),
        disposition=Disposition.OUT_OF_SCOPE,
        reason="Agent-chosen application dependency on the nominated reference app, not a "
        "platform-owned layer. Attributed to the app and listed separately (plan R11).",
    ),
)

#: Anticipated ADDITIONS. These match only against rows that appear in the post-remediation
#: export and were never on the original list — a class the reconcile step otherwise reports as
#: "the fix did not land". `git` is the honest case: the backend shells out to it at runtime and
#: its base ships none, so installing it is a live-defect fix that knowingly ADDS findings
#: during a CVE remediation. Saying so plainly is the point; discovering it in the reviewer's
#: console is not.
ADDITION_RULES: Final[tuple[Rule, ...]] = (
    Rule(
        name="backend-git-runtime-dependency",
        images=(BACKEND,),
        software_names=("git", "git-man", "libcurl", "curl", "ca-certificates"),
        disposition=Disposition.EXCEPTION,
        unit="U1",
        reason="Knowingly-accepted ADDITION. The backend shells out to git at runtime for "
        "snapshot restore and publish, and its base image shipped none — so publish was "
        "broken. Installing git fixes a live production defect and partially offsets the "
        "reduction this pass reports.",
    ),
)


@dataclass(frozen=True, slots=True)
class Verdict:
    """The disposition assigned to a finding, with the reason that defends it."""

    disposition: Disposition
    reason: str
    rule: str
    unit: str = ""
    owner: str = ""
    target_date: str = ""


def disposition_for(f: Finding, *, rules: Sequence[Rule] = RULES) -> Verdict:
    """Assign exactly one disposition to one scanner row.

    DERIVED verdicts are evaluated FIRST and deliberately outrank every structural rule:

    * Already at or past the vendor's fix → DISPUTE. This is true regardless of what we are
      about to change, and conceding it as an exception would concede a vulnerability we do
      not have (plan AE2).
    * Only a pre-release fix exists → EXCEPTION with the pre-release status as its reason.
      Scope Boundaries rules out pre-release language runtimes, so there is no version move
      available and calling it deferred would imply one (plan AE3).

    The final fallback splits on fixability, because those are the two claims a reviewer can
    check independently: a fix exists and we did not take it (DEFERRED, which must carry an
    owner and a target date and is NEVER filed as not-affected), or no fix exists (EXCEPTION).
    """
    if already_at_or_past_fix(f.software_version, f.fixed_version):
        return Verdict(
            disposition=Disposition.DISPUTE,
            reason=(
                f"Installed version {f.software_version} already meets or exceeds the "
                f"vendor's fixed version ({f.fixed_version}) on that release line. Raised "
                "with the scan owner as a suspected scanner error; excluded from the "
                "exception count."
            ),
            rule="derived:already-fixed",
        )
    if f.prerelease_only_fix:
        return Verdict(
            disposition=Disposition.EXCEPTION,
            reason=(
                f"The only fix offered is a pre-release ({f.fixed_version}). The platform "
                "does not ship pre-release language runtimes, so this is accepted risk with "
                "the pre-release status as its stated reason — not a deferral, because there "
                "is no released version to move to."
            ),
            rule="derived:prerelease-only-fix",
        )
    for rule in rules:
        if rule.matches(f):
            return Verdict(
                disposition=rule.disposition,
                reason=rule.reason,
                rule=rule.name,
                unit=rule.unit,
                owner=rule.owner,
                target_date=rule.target_date,
            )
    if f.fixable:
        return Verdict(
            disposition=Disposition.DEFERRED,
            reason=(
                f"A released fix exists ({f.fixed_version}) and was deliberately not taken in "
                "this pass. Needs a named owner and a target date before handover."
            ),
            rule="fallback:fixable",
            owner="UNASSIGNED",
            target_date="UNSET",
        )
    return Verdict(
        disposition=Disposition.EXCEPTION,
        reason="No fixed version is named by the vendor feed for this package.",
        rule="fallback:no-upstream-fix",
    )


# ─────────────────────────────────────────────────────────────────────────────
# Entries — the reporting granularity
# ─────────────────────────────────────────────────────────────────────────────


@dataclass(slots=True)
class Entry:
    """One (image, CVE, package, version) finding, carrying the scanner rows it accounts for."""

    image: str
    cve_id: str
    severity: str
    software_name: str
    software_version: str
    fixed_version: str
    package_manager: str
    software_type: str
    install_paths: list[str] = field(default_factory=list)
    rows: list[Finding] = field(default_factory=list)
    verdict: Verdict = field(
        default_factory=lambda: Verdict(Disposition.HELD, "unassigned", "none")
    )
    vendor_status: str = ""
    tracker_url: str = ""
    replacing_digest: str = ""

    @property
    def rows_accounted(self) -> int:
        return len(self.rows)

    @property
    def key(self) -> tuple[str, str, str, str]:
        return (self.image, self.cve_id, self.software_name, self.software_version)

    @property
    def digest(self) -> str:
        """The manifest digest the scanner measured these rows against, if it named one."""
        for r in self.rows:
            if d := asset_digest(r.asset):
                return d
        return ""

    @property
    def max_epss(self) -> str:
        scores = [r.epss for r in self.rows if r.epss]
        if not scores:
            return ""
        try:
            return f"{max(float(s) for s in scores):.4f}"
        except ValueError:
            return scores[0]

    @property
    def exploited(self) -> str:
        return "True" if any(r.exploited == "True" for r in self.rows) else "False"

    @property
    def first_detected(self) -> str:
        seen = sorted({r.first_detected for r in self.rows if r.first_detected})
        return seen[0] if seen else ""


#: Severity ordering for a stable, reviewer-friendly sort.
_SEVERITY_RANK: Final[dict[str, int]] = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3}


def build_entries(findings: Iterable[Finding], *, rules: Sequence[Rule] = RULES) -> list[Entry]:
    """Collapse scanner rows into entries and assign each one a disposition.

    A group's disposition comes from its rows, which share an image, a CVE, a package and a
    version — so they share every field the structural rules match on except the installation
    path. Where paths disagree the STRICTEST verdict wins, so a finding that is cleared on one
    path and retained on another is never reported as fully cleared.
    """
    grouped: dict[tuple[str, str, str, str], Entry] = {}
    for f in findings:
        entry = grouped.get(f.key)
        if entry is None:
            entry = Entry(
                image=f.image,
                cve_id=f.cve_id,
                severity=f.severity,
                software_name=f.software_name,
                software_version=f.software_version,
                fixed_version=f.fixed_version,
                package_manager=f.package_manager,
                software_type=f.software_type,
            )
            grouped[f.key] = entry
        entry.rows.append(f)
        if f.install_path:
            for part in f.install_path.split(";"):
                p = part.strip()
                if p and p not in entry.install_paths:
                    entry.install_paths.append(p)
        # Keep the most severe label a group's rows carry.
        if _SEVERITY_RANK.get(f.severity, 9) < _SEVERITY_RANK.get(entry.severity, 9):
            entry.severity = f.severity

    for entry in grouped.values():
        verdicts = [disposition_for(r, rules=rules) for r in entry.rows]
        entry.verdict = min(verdicts, key=lambda v: _STRICTNESS[v.disposition])

    return sorted(
        grouped.values(),
        key=lambda e: (
            e.image,
            _SEVERITY_RANK.get(e.severity, 9),
            e.software_name,
            e.cve_id,
        ),
    )


#: Lower is stricter. A row we must still defend outranks one we claim to have cleared, so a
#: mixed group can never be reported as fully Fixed.
_STRICTNESS: Final[dict[Disposition, int]] = {
    Disposition.HELD: 0,
    Disposition.DEFERRED: 1,
    Disposition.EXCEPTION: 2,
    Disposition.DISPUTE_REJECTED: 3,
    Disposition.SUPERSEDED: 4,
    Disposition.DISPUTE: 5,
    Disposition.OUT_OF_SCOPE: 6,
    Disposition.FIXED: 7,
}


# ─────────────────────────────────────────────────────────────────────────────
# Reconciliation
# ─────────────────────────────────────────────────────────────────────────────


@dataclass(slots=True)
class Reconciliation:
    """What the after-export actually says, measured against what the map predicted."""

    entries: list[Entry]
    #: Predicted cleared and genuinely gone from the after-export.
    cleared: int = 0
    #: Predicted cleared but STILL PRESENT — the fix did not land. This is the number that
    #: matters most: it is a broken claim, not a new exception.
    failed_to_clear: list[Entry] = field(default_factory=list)
    #: Present in the after-export and never anticipated by the map.
    unanticipated: list[Entry] = field(default_factory=list)
    #: EVERY entry sourced from the after-export with no before-row — anticipated additions
    #: (installing git) and unanticipated ones alike. The Summary needs this separately from
    #: `unanticipated`: these rows are written to the content sheets but have no counterpart in
    #: the before-export's row total, so a partition check that compares sheets against `before`
    #: alone reads MISMATCH for a perfectly correct report. See `_write_summary`.
    after_only: list[Entry] = field(default_factory=list)
    #: Residual against a manifest we no longer ship — see `reconcile` for why this is not the
    #: same thing as a failed fix, and why conflating them destroys the reduction claim.
    superseded: list[Entry] = field(default_factory=list)


def reconcile(
    before: Sequence[Entry],
    after: Sequence[Entry],
    *,
    addition_rules: Sequence[Rule] = ADDITION_RULES,
    current_digests: dict[str, str] | None = None,
) -> Reconciliation:
    """Measure the after-export against the coverage map.

    Four outcomes, and only one of them is good:

    * A before-entry predicted FIXED that is absent from the after-export — the claim held.
    * A before-entry predicted FIXED that is STILL THERE — the fix did not land. Reporting it
      as a fresh exception would launder a failed remediation into a defended one, so it is
      surfaced under its own heading and its `Fixed` claim is withdrawn.
    * An after-entry with no before-entry — an ADDITION. Anticipated additions (installing git
      to fix the publish defect) carry their stated reason; anything else is reported as
      unanticipated so it is investigated rather than absorbed.
    * An after-entry measured against a manifest we no longer ship — SUPERSEDED.

    THE SUPERSEDED CASE IS THE ONE THAT DECIDES THE REDUCTION CLAIM, and it is invisible
    without `current_digests`. Every artifact here ships under a mutable tag, so a push untags
    the previous manifest without deleting it. If the scanner enumerates ALL retained manifests
    rather than only what a tag points at, the old image keeps reporting its full finding set
    forever — and a naive reconcile reads those rows as "predicted cleared, still present" and
    reports the entire remediation as failed. The fix landed; the scan is looking at the wrong
    object.

    So a residual row is checked against the digest we actually ship BEFORE it is judged. Pass
    `current_digests` as `{image: "sha256:..."}`. Without it this check is skipped rather than
    guessed — and, importantly, its absence is visible in the report rather than silently
    changing every verdict.
    """
    before_by_key = {e.key: e for e in before}
    after_by_key = {e.key: e for e in after}
    shipped = dict(current_digests or {})

    def replacing_digest_for(entry: Entry | None) -> str:
        """The digest we ship, when `entry` was measured against a different one."""
        if entry is None:
            return ""
        current = shipped.get(entry.image, "")
        if not current or not entry.digest or entry.digest == current:
            return ""
        return current

    def mark_superseded(entry: Entry, replacing: str, measured: str) -> None:
        entry.replacing_digest = replacing
        entry.verdict = Verdict(
            disposition=Disposition.SUPERSEDED,
            reason=(
                f"Measured against {measured}, a manifest this platform no longer ships. The "
                f"shipped artifact is {replacing}. Retained manifests keep reporting their "
                "original findings because a push untags rather than deletes — this row "
                "describes an object no deployment runs."
            ),
            rule="reconcile:superseded-artifact",
            unit=entry.verdict.unit,
        )

    result = Reconciliation(entries=[])
    for key, entry in before_by_key.items():
        after_entry = after_by_key.get(key)
        survived = after_entry is not None
        if (replacing := replacing_digest_for(after_entry)) and after_entry is not None:
            mark_superseded(entry, replacing, after_entry.digest)
            result.superseded.append(entry)
        elif entry.verdict.disposition is Disposition.FIXED and survived:
            entry.verdict = Verdict(
                disposition=Disposition.DEFERRED,
                reason=(
                    f"PREDICTED CLEARED BY {entry.verdict.unit or 'this pass'} BUT STILL "
                    f"PRESENT in the post-remediation scan. The change did not reach this "
                    f"package. Original claim: {entry.verdict.reason}"
                ),
                rule="reconcile:failed-to-clear",
                unit=entry.verdict.unit,
                owner="UNASSIGNED",
                target_date="UNSET",
            )
            result.failed_to_clear.append(entry)
        elif entry.verdict.disposition is Disposition.FIXED:
            result.cleared += 1
        elif not survived:
            # Predicted residual but gone anyway — cleared as a side effect. Honest to
            # report as Fixed, with the original prediction preserved in the reason.
            entry.verdict = Verdict(
                disposition=Disposition.FIXED,
                reason=(
                    "Cleared as a side effect of this pass; the coverage map had predicted "
                    f"it would remain ({entry.verdict.disposition}). Original reason: "
                    f"{entry.verdict.reason}"
                ),
                rule="reconcile:cleared-unexpectedly",
                unit=entry.verdict.unit,
            )
            result.cleared += 1
        result.entries.append(entry)

    for key, entry in after_by_key.items():
        if key in before_by_key:
            continue
        result.after_only.append(entry)
        if replacing := replacing_digest_for(entry):
            mark_superseded(entry, replacing, entry.digest)
            result.superseded.append(entry)
            result.entries.append(entry)
            continue
        verdict: Verdict | None = None
        for rule in addition_rules:
            if any(rule.matches(r) for r in entry.rows):
                verdict = Verdict(
                    disposition=rule.disposition,
                    reason=rule.reason,
                    rule=f"addition:{rule.name}",
                    unit=rule.unit,
                )
                break
        if verdict is None:
            verdict = Verdict(
                disposition=Disposition.DEFERRED,
                reason=(
                    "NOT ANTICIPATED by the coverage map. Present in the post-remediation "
                    "scan with no corresponding row on the original list — investigate "
                    "before filing as an exception."
                ),
                rule="reconcile:unanticipated",
                owner="UNASSIGNED",
                target_date="UNSET",
            )
            result.unanticipated.append(entry)
        entry.verdict = verdict
        result.entries.append(entry)

    return result


# ─────────────────────────────────────────────────────────────────────────────
# Human annotations
# ─────────────────────────────────────────────────────────────────────────────


@dataclass(frozen=True, slots=True)
class Override:
    """A human decision the structural rules cannot derive, applied to matching entries.

    Two of the report's required states exist ONLY here, because neither is inferable from a
    scan:

    * **dispute-rejected.** A dispute is a round-trip: we tell the scan owner their finding is
      wrong, and it can come back negative. A rejected dispute must fall back to accepted risk
      WITH a stated reason — never quietly vanish from the report, which is what would happen if
      the only way to record it were deleting the dispute row.
    * **a named owner and a target date on a deferral.** "Affected, fix available, deliberately
      not taken" is only an honest disposition if somebody owns it and there is a date. Without
      those it is indistinguishable from an oversight, and the generator deliberately emits
      `UNASSIGNED`/`UNSET` so an unowned deferral is loud rather than tidy.

    It also carries the vendor's own status text and tracker URL, so every exception row is
    checkable against a public tracker in one click rather than on our say-so.

    Matching is by image + CVE + software name, with an optional version. Anything omitted
    matches every version of that package.
    """

    image: str
    cve_id: str
    software_name: str
    software_version: str = ""
    disposition: Disposition | None = None
    reason: str = ""
    owner: str = ""
    target_date: str = ""
    vendor_status: str = ""
    tracker_url: str = ""

    def matches(self, e: Entry) -> bool:
        return (
            e.image == self.image
            and e.cve_id == self.cve_id
            and e.software_name == self.software_name
            and (not self.software_version or e.software_version == self.software_version)
        )


def load_overrides(path: Path) -> list[Override]:
    """Read the annotations file. Lives in the local working directory, never the repo.

    Its content is per-CVE commentary about a live system — client data by the same argument
    that keeps the workbooks out of the tree (R13). The SHAPE is code and is tested; the
    content is not committed.
    """
    raw: Any = json.loads(path.read_text(encoding="utf-8"))
    rows: Any = raw.get("entries", []) if isinstance(raw, dict) else raw
    out: list[Override] = []
    for i, row in enumerate(rows):
        disposition_text = str(row.get("disposition") or "").strip()
        try:
            disposition = Disposition(disposition_text) if disposition_text else None
        except ValueError:
            valid = ", ".join(sorted(d.value for d in Disposition))
            raise ValueError(
                f"{path}: entry {i} has unknown disposition {disposition_text!r}. "
                f"Valid values: {valid}"
            ) from None
        # A CHANGED DISPOSITION MUST STATE WHY. `apply_overrides` falls back to the previous
        # rule's reason when none is given, so a dispute rejected by the scan owner would ship
        # as accepted risk still carrying the "suspected scanner error; excluded from the
        # exception count" text it was rejected FOR — the reviewer reads a justification that
        # argues against the disposition beside it. Annotation-only overrides (owner, date,
        # vendor status, tracker URL) are unaffected and stay valid without a reason.
        if disposition is not None and not str(row.get("reason") or "").strip():
            raise ValueError(
                f"{path}: entry {i} sets disposition {disposition_text!r} without a `reason`. "
                "A changed disposition must carry the justification it is changed to — "
                "inheriting the previous rule's reason would contradict it."
            )
        # image / cve_id / software_name are REQUIRED: subscript, never `.get`, so a malformed
        # annotation fails here rather than silently matching nothing.
        out.append(
            Override(
                image=str(row["image"]),
                cve_id=str(row["cve_id"]),
                software_name=str(row["software_name"]),
                software_version=str(row.get("software_version") or ""),
                disposition=disposition,
                reason=str(row.get("reason") or ""),
                owner=str(row.get("owner") or ""),
                target_date=str(row.get("target_date") or ""),
                vendor_status=str(row.get("vendor_status") or ""),
                tracker_url=str(row.get("tracker_url") or ""),
            )
        )
    return out


def apply_overrides(entries: Sequence[Entry], overrides: Sequence[Override]) -> list[str]:
    """Apply human annotations. Returns the overrides that matched NOTHING.

    An unmatched override is returned rather than ignored because the realistic mistake is a
    typo in a CVE id or a package name, and the symptom of ignoring it is a report that is
    silently missing the exact row somebody went to the trouble of annotating. The CLI turns a
    non-empty return into a non-zero exit.
    """
    unmatched: list[str] = []
    for ov in overrides:
        hits = [e for e in entries if ov.matches(e)]
        if not hits:
            unmatched.append(f"{ov.image}/{ov.cve_id}/{ov.software_name}")
            continue
        for e in hits:
            e.vendor_status = ov.vendor_status or e.vendor_status
            e.tracker_url = ov.tracker_url or e.tracker_url
            e.verdict = Verdict(
                disposition=ov.disposition or e.verdict.disposition,
                reason=ov.reason or e.verdict.reason,
                rule=f"{e.verdict.rule}+override",
                unit=e.verdict.unit,
                owner=ov.owner or e.verdict.owner,
                target_date=ov.target_date or e.verdict.target_date,
            )
    return unmatched


# ─────────────────────────────────────────────────────────────────────────────
# Workbook
# ─────────────────────────────────────────────────────────────────────────────

COLUMNS: Final[tuple[tuple[str, int], ...]] = (
    ("Image", 26),
    ("CVE ID", 18),
    ("Severity", 10),
    ("Software Name", 24),
    ("Software Version", 24),
    ("Fixed Version", 26),
    ("Package Manager", 16),
    ("Software Type", 14),
    ("Installation Paths", 52),
    ("Rows Accounted", 15),
    ("Disposition", 20),
    ("Reason", 80),
    ("Cleared By / Unit", 16),
    ("Owner", 14),
    ("Target Date", 13),
    ("Vendor Status", 30),
    ("Tracker URL", 34),
    ("Replacing Digest", 24),
    ("Max EPSS", 10),
    ("Exploited in the Wild", 20),
    ("First Detected", 14),
)

_COL: Final[dict[str, str]] = {
    name: get_column_letter(i) for i, (name, _) in enumerate(COLUMNS, start=1)
}

_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TITLE_FONT = Font(bold=True, size=14)
_WARN_FONT = Font(bold=True, color="C00000")


def _row_values(e: Entry) -> list[object]:
    return [
        e.image,
        e.cve_id,
        e.severity,
        e.software_name,
        e.software_version,
        e.fixed_version or "-",
        e.package_manager,
        e.software_type,
        "; ".join(e.install_paths),
        e.rows_accounted,
        str(e.verdict.disposition),
        e.verdict.reason,
        e.verdict.unit,
        e.verdict.owner,
        e.verdict.target_date,
        e.vendor_status,
        e.tracker_url,
        e.replacing_digest,
        e.max_epss,
        e.exploited,
        e.first_detected,
    ]


def _write_content_sheet(wb: Workbook, title: str, entries: Sequence[Entry]) -> None:
    ws = wb.create_sheet(title)
    ws.append([name for name, _ in COLUMNS])
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for e in entries:
        ws.append(_row_values(e))
    for i, (_, width) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"
    if entries:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(entries) + 1}"


def _sumifs(sheet: str, image: str, *, disposition: str = "", severity: str = "") -> str:
    """An Excel SUMIFS over a content sheet's `Rows Accounted` column.

    Summary is DERIVED, never typed: every number on it is a formula over the other sheets, so
    a hand-edit to a content sheet moves the summary with it and the report cannot drift
    internally. That is a property the reviewer can verify by clicking a cell.
    """
    j, a = _COL["Rows Accounted"], _COL["Image"]
    parts = [f"'{sheet}'!${j}:${j}", f"'{sheet}'!${a}:${a}", f'"{image}"']
    if disposition:
        k = _COL["Disposition"]
        parts += [f"'{sheet}'!${k}:${k}", f'"{disposition}"']
    if severity:
        c = _COL["Severity"]
        parts += [f"'{sheet}'!${c}:${c}", f'"{severity}"']
    return f"=SUMIFS({','.join(parts)})"


def _sum_across(sheets: Iterable[str], image: str) -> str:
    j, a = _COL["Rows Accounted"], _COL["Image"]
    terms = [f"SUMIFS('{s}'!${j}:${j},'{s}'!${a}:${a},\"{image}\")" for s in sheets]
    return "=" + "+".join(terms)


def _write_summary(
    wb: Workbook,
    *,
    images: Sequence[str],
    before_rows: dict[str, int],
    after_rows: dict[str, int | None],
    digests: dict[str, str],
    generated_at: str,
    before_files: Sequence[str],
    after_files: Sequence[str],
    held: int,
    failed_to_clear: int,
    unanticipated: int,
    addition_rows: dict[str, int] | None = None,
    never_scanned: Sequence[str] = (),
) -> None:
    addition_rows = addition_rows or {}
    ws = wb.create_sheet("Summary", 0)
    ws["A1"] = "BIAL container image remediation — what was fixed, what was not, and why"
    ws["A1"].font = _TITLE_FONT
    ws["A2"] = f"Generated {generated_at} · input: {', '.join(before_files) or '—'}"
    ws["A3"] = (
        f"Post-remediation export: {', '.join(after_files)}"
        if after_files
        else "Post-remediation export: NOT YET RECEIVED — this is the coverage map, not the "
        "final report. The `after` and `reduction` columns are unfilled by design."
    )
    ws["A4"] = (
        "Every number below is a FORMULA over the other sheets — never typed. Each row of the "
        "original scan appears in exactly one sheet, and Rows Accounted sums back to the "
        "scanner's own row totals, so this report is arithmetically self-checking."
    )
    ws["A5"] = (
        "Granularity: one entry per image, CVE, package name and version, carrying the number "
        "of scanner rows it accounts for. BIAL's scanner emits one row per affected sibling "
        "package and per installation path."
    )

    row = 7
    if held:
        ws.cell(row=row, column=1, value=f"DRAFT — {held} scanner rows are HELD pending an answer")
        ws.cell(row=row, column=1).font = _WARN_FONT
        ws.cell(row=row, column=2, value="These must be dispositioned before this report ships.")
        row += 1
    if failed_to_clear:
        ws.cell(row=row, column=1, value=f"{failed_to_clear} entries FAILED TO CLEAR")
        ws.cell(row=row, column=1).font = _WARN_FONT
        ws.cell(
            row=row,
            column=2,
            value="Predicted cleared by this pass but still present in the rescan — the change "
            "did not reach them. Listed on `Deferred`.",
        )
        row += 1
    if unanticipated:
        ws.cell(row=row, column=1, value=f"{unanticipated} entries were NOT ANTICIPATED")
        ws.cell(row=row, column=1).font = _WARN_FONT
        ws.cell(
            row=row,
            column=2,
            value="Present in the rescan with no row on the original list. Investigate before "
            "filing as exceptions.",
        )
        row += 1
    if held or failed_to_clear or unanticipated:
        row += 1

    # THE PARTITION CHECK COMPARES LIKE WITH LIKE, which needs the `Additions` column to exist.
    # The content sheets hold two populations: rows that came from the BEFORE export, and rows
    # the after-export introduced with no before-row (the anticipated `git` install is the whole
    # reason `ADDITION_RULES` exists). Checking `Sheets total = Before rows` therefore reported
    # MISMATCH on the backend row and on the grand TOTAL for a report that was entirely correct
    # — and it did so while the CLI printed `partition : OK`, because the Python-side check runs
    # over the before-export alone. A self-check that cries wolf on the one case the tool was
    # built to handle is worse than no self-check: the reviewer learns to ignore the column.
    headers = [
        "Image",
        "Scanned digest (before)",
        "Before rows",
        "After rows",
        "Reduction",
        "Fixed",
        "Exceptions",
        "Deferred",
        "Disputes",
        "Out of scope",
        "Held",
        "Additions (after only)",
        "Sheets total",
        "Partition check",
    ]
    header_row = row
    for c, name in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    row += 1

    never = frozenset(never_scanned)
    first_data_row = row
    for image in images:
        after = after_rows.get(image)
        additions = addition_rows.get(image, 0)
        ws.cell(row=row, column=1, value=image)
        ws.cell(row=row, column=2, value=digests.get(image, "") or "not scanned")
        # An image with no before-scan says so. A typed 0 would read as "scanned, found nothing",
        # which is the opposite claim and the flattering one.
        if image in never:
            ws.cell(row=row, column=3, value="not scanned")
        else:
            ws.cell(row=row, column=3, value=before_rows.get(image, 0))
        if after is None:
            ws.cell(row=row, column=4, value="never scanned")
            ws.cell(row=row, column=5, value="n/a — after only")
        elif image in never:
            ws.cell(row=row, column=4, value=after)
            ws.cell(row=row, column=5, value="n/a — after only")
        else:
            ws.cell(row=row, column=4, value=after)
            ws.cell(row=row, column=5, value=f"=C{row}-D{row}")
        ws.cell(row=row, column=6, value=_sumifs("Fixed", image))
        ws.cell(row=row, column=7, value=_sumifs("Exceptions", image))
        ws.cell(row=row, column=8, value=_sumifs("Deferred", image))
        ws.cell(row=row, column=9, value=_sumifs("Disputes", image))
        ws.cell(row=row, column=10, value=_sumifs("Out of scope", image))
        ws.cell(row=row, column=11, value=_sumifs("Held", image))
        ws.cell(row=row, column=12, value=additions)
        ws.cell(row=row, column=13, value=_sum_across(CONTENT_SHEETS, image))
        # A never-scanned image has no before total to add to, so its sheets must equal its
        # additions alone. Adding the text "not scanned" would yield #VALUE!.
        if image in never:
            ws.cell(row=row, column=14, value=f'=IF(M{row}=L{row},"OK","MISMATCH")')
        else:
            ws.cell(row=row, column=14, value=f'=IF(M{row}=C{row}+L{row},"OK","MISMATCH")')
        row += 1

    last_data_row = row - 1
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    for c in range(3, 14):
        col = get_column_letter(c)
        if c in (4, 5) and any(v is None for v in after_rows.values()):
            ws.cell(row=row, column=c, value="—")
            continue
        # SUM ignores the text cells a never-scanned row puts in C, so this stays the numeric
        # before-total rather than erroring.
        ws.cell(row=row, column=c, value=f"=SUM({col}{first_data_row}:{col}{last_data_row})")
        ws.cell(row=row, column=c).font = Font(bold=True)
    ws.cell(row=row, column=14, value=f'=IF(M{row}=C{row}+L{row},"OK","MISMATCH")').font = Font(
        bold=True
    )
    row += 2

    ws.cell(row=row, column=1, value="Severity breakdown (scanner rows, before)").font = Font(
        bold=True
    )
    row += 1
    sev_headers = ["Image", *_SEVERITY_RANK.keys(), "Total"]
    for c, name in enumerate(sev_headers, start=1):
        cell = ws.cell(row=row, column=c, value=name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
    row += 1
    for image in images:
        ws.cell(row=row, column=1, value=image)
        for c, sev in enumerate(_SEVERITY_RANK, start=2):
            terms = [
                f"SUMIFS('{s}'!${_COL['Rows Accounted']}:${_COL['Rows Accounted']},"
                f"'{s}'!${_COL['Image']}:${_COL['Image']},\"{image}\","
                f"'{s}'!${_COL['Severity']}:${_COL['Severity']},\"{sev}\")"
                for s in CONTENT_SHEETS
            ]
            ws.cell(row=row, column=c, value="=" + "+".join(terms))
        ws.cell(row=row, column=6, value=f"=SUM(B{row}:E{row})")
        row += 1

    widths = [26, 74, 12, 14, 12, 10, 12, 11, 11, 13, 8, 13, 13, 15]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{header_row + 1}"


def build_workbook(
    entries: Sequence[Entry],
    *,
    before_rows: dict[str, int],
    after_rows: dict[str, int | None],
    digests: dict[str, str],
    generated_at: str,
    before_files: Sequence[str] = (),
    after_files: Sequence[str] = (),
    failed_to_clear: int = 0,
    unanticipated: int = 0,
    addition_rows: dict[str, int] | None = None,
    never_scanned: Sequence[str] = (),
) -> Workbook:
    """Build the report workbook. `Summary` first, then one sheet per disposition group.

    `addition_rows` carries the scanner rows that entered from the AFTER export with no
    before-row, per image. The Summary's partition check needs them as their own term — see
    `_write_summary` — because the content sheets hold both populations and comparing their
    total against the before-export alone reports MISMATCH for a correct report.
    """
    wb = Workbook()
    if (default_sheet := wb.active) is not None:
        wb.remove(default_sheet)  # drop openpyxl's default sheet

    by_sheet: dict[str, list[Entry]] = {name: [] for name in CONTENT_SHEETS}
    for e in entries:
        by_sheet[SHEET_FOR[e.verdict.disposition]].append(e)

    images = sorted({*before_rows, *(k for k in after_rows)})
    held_rows = sum(e.rows_accounted for e in by_sheet["Held"])
    _write_summary(
        wb,
        images=images,
        before_rows=before_rows,
        after_rows=after_rows,
        digests=digests,
        generated_at=generated_at,
        before_files=before_files,
        after_files=after_files,
        held=held_rows,
        failed_to_clear=failed_to_clear,
        unanticipated=unanticipated,
        addition_rows=addition_rows,
        never_scanned=never_scanned,
    )
    for name in CONTENT_SHEETS:
        _write_content_sheet(wb, name, by_sheet[name])
    return wb


# ─────────────────────────────────────────────────────────────────────────────
# Integrity checks — the report's own self-check
# ─────────────────────────────────────────────────────────────────────────────


@dataclass(slots=True)
class Integrity:
    """The partition invariant, checked in Python so a broken report fails the run."""

    ok: bool
    problems: list[str] = field(default_factory=list)
    total_rows_in: int = 0
    total_rows_out: int = 0
    duplicate_keys: list[str] = field(default_factory=list)
    unmapped: list[str] = field(default_factory=list)


def check_integrity(findings: Sequence[Finding], entries: Sequence[Entry]) -> Integrity:
    """Every scanner row lands in exactly one entry, and every entry carries a disposition.

    A row appearing twice, or in no sheet, is a real error — that is what makes the arithmetic
    a check rather than decoration. This runs BEFORE the workbook is written so a report that
    does not add up is never produced in the first place.
    """
    problems: list[str] = []
    seen_keys: Counter[tuple[str, str, str, str]] = Counter(e.key for e in entries)
    duplicates = [str(k) for k, n in seen_keys.items() if n > 1]
    if duplicates:
        problems.append(f"{len(duplicates)} entry keys appear more than once")

    total_in = len(findings)
    total_out = sum(e.rows_accounted for e in entries)
    if total_in != total_out:
        problems.append(
            f"scanner rows in ({total_in}) != rows accounted for by entries ({total_out})"
        )

    sources_in = Counter(f.source for f in findings)
    sources_out: Counter[str] = Counter()
    for e in entries:
        for r in e.rows:
            sources_out[r.source] += 1
    missing = [s for s in sources_in if s not in sources_out]
    doubled = [s for s, n in sources_out.items() if n > 1]
    if missing:
        problems.append(f"{len(missing)} scanner rows landed in no entry (first: {missing[0]})")
    if doubled:
        problems.append(f"{len(doubled)} scanner rows landed in more than one entry")

    unmapped = [
        f"{e.image}/{e.cve_id}/{e.software_name}"
        for e in entries
        if e.verdict.disposition is Disposition.HELD
    ]

    return Integrity(
        ok=not problems,
        problems=problems,
        total_rows_in=total_in,
        total_rows_out=total_out,
        duplicate_keys=duplicates,
        unmapped=unmapped,
    )


# ─────────────────────────────────────────────────────────────────────────────
# Coverage map output (plan U9)
# ─────────────────────────────────────────────────────────────────────────────

_MAP_COLUMNS: Final[tuple[str, ...]] = (
    "source",
    "image",
    "asset_digest",
    "cve_id",
    "severity",
    "software_name",
    "software_version",
    "fixed_version",
    "fixable",
    "package_manager",
    "software_type",
    "install_path",
    "disposition",
    "rule",
    "unit",
    "reason",
)


def write_coverage_map(
    findings: Sequence[Finding],
    entries: Sequence[Entry],
    out_dir: Path,
    *,
    generated_at: str,
) -> dict[str, Any]:
    """Write the row-level coverage map and its reconciliation summary.

    ROW level, not entry level, on purpose: the map's job is to prove that every row on BIAL's
    list carries exactly one intended disposition, and only a row-level artifact can be checked
    against the reviewer's own console line by line.
    """
    out_dir.mkdir(parents=True, exist_ok=True)
    verdict_by_key = {e.key: e.verdict for e in entries}

    csv_path = out_dir / "coverage-map.csv"
    with csv_path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=list(_MAP_COLUMNS))
        writer.writeheader()
        for f in findings:
            v = verdict_by_key[f.key]
            writer.writerow(
                {
                    "source": f.source,
                    "image": f.image,
                    "asset_digest": asset_digest(f.asset),
                    "cve_id": f.cve_id,
                    "severity": f.severity,
                    "software_name": f.software_name,
                    "software_version": f.software_version,
                    "fixed_version": f.fixed_version or "-",
                    "fixable": "yes" if f.fixable else "no",
                    "package_manager": f.package_manager,
                    "software_type": f.software_type,
                    "install_path": f.install_path,
                    "disposition": str(v.disposition),
                    "rule": v.rule,
                    "unit": v.unit,
                    "reason": v.reason,
                }
            )

    per_image: dict[str, dict[str, Any]] = defaultdict(
        lambda: {"rows": 0, "entries": 0, "distinct_cves": set(), "dispositions": Counter()}
    )
    for f in findings:
        bucket = per_image[f.image]
        bucket["rows"] += 1
        bucket["distinct_cves"].add(f.cve_id)
        bucket["dispositions"][str(verdict_by_key[f.key].disposition)] += 1
    for e in entries:
        per_image[e.image]["entries"] += 1

    summary: dict[str, Any] = {
        "generated_at": generated_at,
        "totals": {
            "scanner_rows": len(findings),
            "entries": len(entries),
            "distinct_cves_global": len({f.cve_id for f in findings}),
        },
        "images": {
            image: {
                "rows": b["rows"],
                "entries": b["entries"],
                "distinct_cves": len(b["distinct_cves"]),
                "digest": next(
                    (asset_digest(f.asset) for f in findings if f.image == image and f.asset), ""
                ),
                "dispositions": dict(sorted(b["dispositions"].items())),
            }
            for image, b in sorted(per_image.items())
        },
    }
    (out_dir / "coverage-summary.json").write_text(
        json.dumps(summary, indent=2, sort_keys=False) + "\n", encoding="utf-8"
    )
    return summary


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────


def _rows_per_image(findings: Iterable[Finding]) -> dict[str, int]:
    counts: Counter[str] = Counter(f.image for f in findings)
    return dict(counts)


def _digests(findings: Iterable[Finding]) -> dict[str, str]:
    out: dict[str, str] = {}
    for f in findings:
        if f.image not in out and (d := asset_digest(f.asset)):
            out[f.image] = d
    return out


def _report(integrity: Integrity, stream: TextIO = sys.stdout) -> None:
    print(f"  scanner rows in : {integrity.total_rows_in}", file=stream)
    print(f"  rows accounted  : {integrity.total_rows_out}", file=stream)
    if integrity.ok:
        print("  partition       : OK — every row lands in exactly one entry", file=stream)
    else:
        for p in integrity.problems:
            print(f"  PARTITION ERROR : {p}", file=stream)


def main(argv: Sequence[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        prog="exception_register",
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    sub = parser.add_subparsers(dest="command", required=True)

    m = sub.add_parser("map", help="pass 1 — build the coverage map from the BEFORE export")
    m.add_argument("--before", action="append", required=True, type=Path)
    m.add_argument("--out-dir", required=True, type=Path)

    r = sub.add_parser("register", help="pass 2 — build the register from BEFORE + AFTER exports")
    r.add_argument("--before", action="append", required=True, type=Path)
    # REQUIRED, and the requirement is load-bearing rather than tidy. `reconcile` reads "absent
    # from the after-export" as "cleared", so an empty after-set re-dispositions EVERY row —
    # exceptions, deferrals, held rows alike — to `fixed` with the reason "cleared as a side
    # effect of this pass", and the tool then reports 100% remediation against data it has never
    # seen. The coverage map is what an operator wants before the rescan lands; that is the `map`
    # subcommand, and it is what this now points them at.
    r.add_argument("--after", action="append", required=True, type=Path)
    r.add_argument("--out", required=True, type=Path)
    r.add_argument(
        "--never-scanned",
        action="append",
        default=[],
        help="an image that has no BEFORE scan (its Summary row carries an after only)",
    )
    r.add_argument(
        "--current-digest",
        action="append",
        default=[],
        metavar="IMAGE=sha256:...",
        help=(
            "the manifest this platform actually ships for IMAGE. Residual rows measured "
            "against any other digest are reported as superseded-artifact rather than as a "
            "failed fix — without this, a scanner that enumerates retained manifests makes a "
            "successful remediation read as a total failure"
        ),
    )
    r.add_argument(
        "--overrides",
        type=Path,
        help=(
            "JSON annotations the scan cannot supply: rejected disputes, and the owner + "
            "target date every deferral needs. Lives in the local working directory"
        ),
    )

    args = parser.parse_args(argv)
    generated_at = datetime.now(UTC).strftime("%Y-%m-%d %H:%M UTC")

    before_findings = load_exports(args.before)
    if not before_findings:
        print("No findings parsed from the BEFORE export(s).", file=sys.stderr)
        return 2
    before_entries = build_entries(before_findings)
    integrity = check_integrity(before_findings, before_entries)

    if args.command == "map":
        summary = write_coverage_map(
            before_findings, before_entries, args.out_dir, generated_at=generated_at
        )
        print(f"Coverage map written to {args.out_dir}")
        for image, b in summary["images"].items():
            print(f"  {image:24s} rows={b['rows']:5d} entries={b['entries']:5d} {b['digest']}")
            for d, n in b["dispositions"].items():
                print(f"      {d:22s} {n:5d}")
        _report(integrity)
        if integrity.unmapped:
            print(f"  HELD            : {len(integrity.unmapped)} entries await an answer")
        return 0 if integrity.ok else 1

    after_findings = load_exports(args.after)
    if not after_findings:
        print(
            "No findings parsed from the AFTER export(s). `register` reconciles a "
            "post-remediation scan against the coverage map; with nothing to reconcile it "
            "would report every finding as cleared. Use the `map` subcommand for the "
            "pre-rescan coverage map.",
            file=sys.stderr,
        )
        return 2
    after_entries = build_entries(after_findings)

    current_digests: dict[str, str] = {}
    for pair in args.current_digest:
        image, sep, digest = str(pair).partition("=")
        if not sep or not digest.startswith("sha256:"):
            print(
                f"--current-digest must be IMAGE=sha256:... — got {pair!r}",
                file=sys.stderr,
            )
            return 2
        current_digests[image.strip()] = digest.strip()

    # THE DIGEST MUST BE ONE THE SCAN ACTUALLY SAW. `reconcile` treats "residual row measured
    # against a digest other than the one we ship" as superseded-artifact and drops it out of the
    # reduction — correct when the scanner enumerates retained manifests, catastrophic when the
    # operator simply mistyped or pasted the wrong one. A digest that matches NO row for an image
    # the scan does cover marks that image's ENTIRE residual set superseded, laundering a total
    # remediation failure into a clean report with a zero exit code.
    #
    # The realistic mistake is not a typo: it is pasting the INDEX digest where the scanner
    # reports the per-architecture child (or the reverse), which is exactly the distinction this
    # remediation's own Dockerfiles spend paragraphs on.
    after_digests_by_image: dict[str, set[str]] = defaultdict(set)
    for f in after_findings:
        if d := asset_digest(f.asset):
            after_digests_by_image[f.image].add(d)
    for image, digest in current_digests.items():
        seen = after_digests_by_image.get(image, set())
        if seen and digest not in seen:
            print(
                f"--current-digest {image}={digest} matches no manifest in the AFTER export "
                f"(it names {', '.join(sorted(seen))}). Every residual row for this image "
                "would be filed as superseded-artifact and dropped from the reduction. Check "
                "you are not passing an index digest where the scan reports the "
                "per-architecture child.",
                file=sys.stderr,
            )
            return 2

    rec = reconcile(before_entries, after_entries, current_digests=current_digests)

    unmatched: list[str] = []
    if args.overrides:
        unmatched = apply_overrides(rec.entries, load_overrides(args.overrides))

    before_rows = _rows_per_image(before_findings)
    after_rows: dict[str, int | None] = {
        image: count for image, count in _rows_per_image(after_findings).items()
    }
    for image in args.never_scanned:
        before_rows.setdefault(image, 0)
        after_rows[image] = after_rows.get(image)

    # AN IMAGE THE AFTER-EXPORT NEVER MENTIONS IS NOT A REMEDIATED IMAGE. Filling a missing
    # after-count with a typed 0 renders as a clean sweep: full reduction for that image, its HELD
    # rows rewritten to fixed, and the `held_now` ship-gate satisfied precisely because the rows it
    # guards stopped existing. A STAGGERED RESCAN — BIAL returns two of the three images this week
    # and the third next week — is the ordinary case, not an exotic one, so the failure fires on a
    # normal Tuesday and hands the client a report claiming an image was cleared that nobody
    # looked at. That is the same false assurance the `--after` guard above exists to prevent, one
    # level up: there the whole export was empty, here one image of it is.
    #
    # `--never-scanned` already carries the honest encoding for "no data here" — it stores None,
    # which the workbook renders as text rather than a numeric zero, and which the loop above has
    # already keyed. So the operator has a way to say it; they just have to say it out loud.
    missing_after = sorted(set(before_rows) - set(after_rows))
    if missing_after:
        print(
            "these images appear in the BEFORE export but in no AFTER export: "
            f"{', '.join(missing_after)}. Their residual rows cannot be measured, and recording "
            "them as zero would report a full reduction for an image that was never rescanned. "
            "Pass the missing after-export, or declare the image with --never-scanned to record "
            'it honestly as "not scanned".',
            file=sys.stderr,
        )
        return 2

    addition_rows: dict[str, int] = {}
    for entry in rec.after_only:
        addition_rows[entry.image] = addition_rows.get(entry.image, 0) + entry.rows_accounted

    wb = build_workbook(
        rec.entries,
        before_rows=before_rows,
        after_rows=after_rows,
        digests=_digests(before_findings) | _digests(after_findings),
        generated_at=generated_at,
        before_files=[p.name for p in args.before],
        after_files=[p.name for p in args.after],
        failed_to_clear=len(rec.failed_to_clear),
        unanticipated=len(rec.unanticipated),
        addition_rows=addition_rows,
        never_scanned=args.never_scanned,
    )
    args.out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.out)
    print(f"Register written to {args.out}")
    _report(integrity)
    if rec.superseded:
        print(f"  SUPERSEDED      : {len(rec.superseded)} entries against an unshipped manifest")
    if rec.failed_to_clear:
        print(f"  FAILED TO CLEAR : {len(rec.failed_to_clear)} entries — the fix did not land")
    if rec.unanticipated:
        print(f"  UNANTICIPATED   : {len(rec.unanticipated)} entries absent from the coverage map")
    # THE HELD GATE READS THE WORKBOOK THAT WAS WRITTEN, not the coverage map it started from.
    # `integrity` is computed over `before_entries` BEFORE `reconcile` mutates those same objects
    # and `apply_overrides` annotates them, so `integrity.unmapped` is a snapshot of a report
    # that no longer exists: it blocked a run whose Held sheet had emptied (once any row had ever
    # been HELD, no run could exit 0 again) and would have passed one whose Held sheet had since
    # filled. `integrity`'s partition arithmetic is still correct and still reported — that part
    # is genuinely about the before-export.
    held_now = [
        f"{e.image}/{e.cve_id}/{e.software_name}"
        for e in rec.entries
        if e.verdict.disposition is Disposition.HELD
    ]
    if held_now:
        print(f"  HELD            : {len(held_now)} entries — NOT READY TO SHIP")
    for miss in unmatched:
        print(f"  OVERRIDE MATCHED NOTHING: {miss}", file=sys.stderr)
    unowned = [
        e
        for e in rec.entries
        if e.verdict.disposition is Disposition.DEFERRED
        and (e.verdict.owner in ("", "UNASSIGNED") or e.verdict.target_date in ("", "UNSET"))
    ]
    if unowned:
        print(
            f"  UNOWNED DEFERRAL: {len(unowned)} entries lack an owner or a target date — "
            "annotate them via --overrides before this ships",
            file=sys.stderr,
        )
    ready = integrity.ok and not held_now and not unmatched and not unowned
    return 0 if ready else 1


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(main())
