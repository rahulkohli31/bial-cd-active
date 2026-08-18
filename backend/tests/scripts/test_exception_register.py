"""The remediation report has to be arithmetically self-checking, so its arithmetic is tested.

The report's whole claim to credibility is that a reviewer can ADD IT UP: every row on BIAL's
original list lands in exactly one sheet, the sheets sum back to the scanner's own totals, and
`Summary` is derived from the other sheets rather than typed. Those are properties, not
formatting, and each one has a test here that fails when the property breaks.

NO REAL CLIENT DATA. Every fixture below is synthetic — invented CVE ids, invented packages, a
made-up registry. The real exports are client vulnerability data about a live system and never
enter this repository (the generator's own module docstring explains the split). Synthetic data
is also strictly better for these tests: it lets a fixture encode the exact shape a scenario
needs, including shapes the real export does not currently contain.
"""

from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl
import pytest
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from scripts.exception_register import (
    Disposition,
    Entry,
    Override,
    already_at_or_past_fix,
    apply_overrides,
    build_entries,
    build_workbook,
    check_integrity,
    compare_versions,
    disposition_for,
    is_prerelease,
    load_export,
    load_exports,
    load_overrides,
    main,
    reconcile,
    write_coverage_map,
)

REGISTRY = "testregistry.azurecr.io"
SANDBOX = "citizen-dev-sandbox"
FRONTEND = "vibe-coding-frontend"
BACKEND = "vibe-coding-backend"

DIGEST_OLD = "sha256:" + "a" * 64
DIGEST_NEW = "sha256:" + "b" * 64


# ─────────────────────────────────────────────────────────────────────────────
# Factories — the two export shapes BIAL actually sends
# ─────────────────────────────────────────────────────────────────────────────

_ASSET_HEADERS = [
    "CVE ID",
    "Description",
    "Severity",
    "Status",
    "Asset Name",
    "Software Name",
    "Software Version",
    "Fixed Version",
    "First Detected",
    "Software Package Manager",
    "Exploited in the Wild",
]

_SEVERITY_HEADERS = [
    "CVE ID",
    "Description",
    "Status",
    "Software Name",
    "Software Version",
    "Fixed Version",
    "First Detected",
    "CVE EPSS Score",
    "Software Type",
    "Software Package Manager",
    "Installation Path",
    "Exploited in the Wild",
]


def row(
    cve: str,
    software: str,
    version: str,
    *,
    fixed: str = "-",
    severity: str = "High",
    manager: str = "deb",
    kind: str = "OS",
    path: str = "/var/lib/dpkg/status",
    epss: str = "0.0010",
) -> dict[str, str]:
    """One scanner row, in the shape both writers below understand."""
    return {
        "cve": cve,
        "software": software,
        "version": version,
        "fixed": fixed,
        "severity": severity,
        "manager": manager,
        "kind": kind,
        "path": path,
        "epss": epss,
    }


def write_asset_shape(path: Path, sheets: dict[str, tuple[str, list[dict[str, str]]]]) -> Path:
    """The `vibe-coding_sheet.xlsx` shape: one sheet per ASSET, severity as a column."""
    wb = Workbook()
    if (default := wb.active) is not None:
        wb.remove(default)
    summary = wb.create_sheet("Summary")  # a pivot sheet, which the loader must skip
    summary["A1"] = "Count of CVE ID"
    summary["A2"] = "Row Labels"
    for sheet_name, (asset, rows) in sheets.items():
        ws = wb.create_sheet(sheet_name)
        ws.append(_ASSET_HEADERS)
        for r in rows:
            ws.append(
                [
                    r["cve"],
                    f"synthetic description for {r['cve']}",
                    r["severity"],
                    "New",
                    asset,
                    r["software"],
                    r["version"],
                    r["fixed"],
                    "2026-07-14",
                    r["manager"],
                    "False",
                ]
            )
    wb.save(path)
    return path


def write_severity_shape(path: Path, asset: str, rows: list[dict[str, str]]) -> Path:
    """The `citizen-dev-sandbox` shape: one sheet per SEVERITY, asset only on the pivot."""
    wb = Workbook()
    if (default := wb.active) is not None:
        wb.remove(default)
    summary = wb.create_sheet("Summary")
    summary["A2"] = "CVE COUNT"
    summary["A4"] = asset  # the ONLY place the asset (and its digest) appears
    for severity in ("Critical", "High", "Medium", "Low"):
        ws = wb.create_sheet(severity)
        ws.append(_SEVERITY_HEADERS)
        for r in (x for x in rows if x["severity"] == severity):
            ws.append(
                [
                    r["cve"],
                    f"synthetic description for {r['cve']}",
                    "New",
                    r["software"],
                    r["version"],
                    r["fixed"],
                    "2026-07-23",
                    r["epss"],
                    r["kind"],
                    r["manager"],
                    r["path"],
                    "False",
                ]
            )
    wb.save(path)
    return path


def asset_ref(repo: str, digest: str = DIGEST_OLD) -> str:
    return f"{REGISTRY}/{repo}@{digest}"


# ─────────────────────────────────────────────────────────────────────────────
# Version comparison — the logic a wrong answer here quietly corrupts everything
# ─────────────────────────────────────────────────────────────────────────────


@pytest.mark.parametrize(
    ("a", "b", "expected"),
    [
        # dpkg basics
        ("1.0", "1.0", 0),
        ("1.0", "1.1", -1),
        ("1.10", "1.9", 1),  # numeric, not lexical — the classic bug
        ("2.5.0-1+deb12u2", "2.5.0-1+deb12u3", -1),
        ("7.88.1-10+deb12u15", "7.88.1-10+deb12u9", 1),  # 15 > 9, not "1" < "9"
        # epochs win outright
        ("1:1.0", "2.0", 1),
        # alpine revisions
        ("3.3.3-r0", "3.3.7-r0", -1),
        ("2.13.4-r5", "2.13.4-r10", -1),
        ("1.2.5-r9", "1.2.5-r9", 0),
        # `~` sorts BEFORE the empty string — this is what makes a pre-release older
        ("1.0~rc1", "1.0", -1),
        ("252.39-1~deb12u2", "252.39-1", -1),
        # go/v prefixes are stripped so feeds can be compared with installed versions
        ("go1.22.3", "1.22.4", -1),
        ("v2.8.4", "2.11.4", -1),
        ("go1.24.13", "1.24.13", 0),
    ],
)
def test_compare_versions(a: str, b: str, expected: int) -> None:
    assert compare_versions(a, b) == expected
    assert compare_versions(b, a) == -expected


@pytest.mark.parametrize(
    ("version", "expected"),
    [
        ("3.15.0a6", True),
        ("3.15.0b3", True),
        ("1.26.0-rc.3", True),
        ("2.0.0-beta", True),
        ("1.0~rc1", True),
        # NOT pre-releases — the two shapes a naive regex mangles
        ("2.13.4-r5", False),  # alpine revision
        ("7.88.1-10+deb12u15", False),  # debian point release
        ("3.14.7", False),
        ("go1.23.12", False),
        ("16.2.12", False),
    ],
)
def test_is_prerelease(version: str, expected: bool) -> None:
    assert is_prerelease(version) is expected


@pytest.mark.parametrize(
    ("installed", "fixed", "expected"),
    [
        # THE RELEASE-LINE RULE. Getting this wrong marks nearly every genuine finding as a
        # dispute, which is the one claim a reviewer can disprove from their own console.
        ("go1.23.12", "1.24.9, 1.25.3", False),  # no fix on OUR line — still affected
        ("go1.22.3", "1.21.11, 1.22.4", False),  # same-line fix is 1.22.4 — still affected
        ("go1.22.5", "1.21.11, 1.22.4", True),  # past the same-line fix — dispute
        ("go1.24.13", "1.24.13, 1.25.7", True),  # exactly at the fix — dispute
        # A same-line fix must not be masked by a higher out-of-line candidate.
        ("1.22.9", "1.22.4, 9.9.9", True),
        # No same-line candidate at all: only "past ALL of them" counts.
        ("4.0.0", "2.9, 3.1", True),
        ("2.8", "2.9, 3.1", False),
        # No fix named at all is never a dispute.
        ("1.0", "-", False),
        ("1.0", "", False),
    ],
)
def test_already_at_or_past_fix(installed: str, fixed: str, expected: bool) -> None:
    assert already_at_or_past_fix(installed, fixed) is expected


# ─────────────────────────────────────────────────────────────────────────────
# Ingest
# ─────────────────────────────────────────────────────────────────────────────


def test_both_export_shapes_normalise_to_one_table(tmp_path: Path) -> None:
    """The two workbooks have different columns; the UNION is the schema, not the intersection.

    Dropping the columns only one shape carries would discard the installation path, which is
    the only thing separating the sandbox's vendored-build-binary findings from its OS ones.
    """
    a = write_asset_shape(
        tmp_path / "assets.xlsx",
        {
            FRONTEND: (
                asset_ref(FRONTEND),
                [row("CVE-9000-0001", "libssl3", "3.3.3-r0", manager="apk")],
            ),
            BACKEND: (asset_ref(BACKEND), [row("CVE-9000-0002", "libc6", "2.41-12")]),
        },
    )
    b = write_severity_shape(
        tmp_path / "severity.xlsx",
        asset_ref(SANDBOX),
        [
            row(
                "CVE-9000-0003",
                "stdlib",
                "go1.20.7",
                manager="go-module",
                path="/usr/local/bin/caddy",
            )
        ],
    )

    findings = load_exports([a, b])

    assert len(findings) == 3
    by_image = {f.image: f for f in findings}
    assert set(by_image) == {FRONTEND, BACKEND, SANDBOX}
    # Severity comes from the SHEET NAME in the second shape and from a COLUMN in the first.
    assert by_image[SANDBOX].severity == "High"
    assert by_image[FRONTEND].severity == "High"
    # The install path survives from the shape that has it, and is empty (never invented) in
    # the shape that does not.
    assert by_image[SANDBOX].install_path == "/usr/local/bin/caddy"
    assert by_image[FRONTEND].install_path == ""
    # Provenance is per-row, which is what makes the partition check possible at all.
    assert len({f.source for f in findings}) == 3


def test_the_asset_digest_is_recovered_from_the_pivot_only_shape(tmp_path: Path) -> None:
    """The sandbox export names its asset ONLY on the pivot sheet — losing it loses the digest.

    That digest is the rollback anchor for an artifact shipping under a mutable tag, so a
    loader that skipped the pivot sheet would throw away the one recoverable record of what
    was running.
    """
    path = write_severity_shape(
        tmp_path / "s.xlsx", asset_ref(SANDBOX), [row("CVE-9000-0004", "libexpat1", "2.5.0-1")]
    )
    findings = load_export(path)
    assert len(findings) == 1
    assert findings[0].image == SANDBOX
    assert findings[0].asset.endswith(DIGEST_OLD)


def test_pivot_sheets_are_skipped_not_counted_twice(tmp_path: Path) -> None:
    """A Summary sheet is a VIEW of the rows. Counting it would double every total."""
    path = write_asset_shape(
        tmp_path / "a.xlsx",
        {
            FRONTEND: (
                asset_ref(FRONTEND),
                [row("CVE-9000-0005", "nginx", "1.27.5-r1", manager="apk")],
            )
        },
    )
    assert len(load_export(path)) == 1


# ─────────────────────────────────────────────────────────────────────────────
# Dispositions
# ─────────────────────────────────────────────────────────────────────────────


def _one(tmp_path: Path, image: str, r: dict[str, str], *, name: str = "x.xlsx") -> Entry:
    if image == SANDBOX:
        path = write_severity_shape(tmp_path / name, asset_ref(SANDBOX), [r])
    else:
        path = write_asset_shape(tmp_path / name, {image: (asset_ref(image), [r])})
    entries = build_entries(load_export(path))
    assert len(entries) == 1
    return entries[0]


def test_a_finding_already_past_its_fix_is_a_dispute_not_an_exception(tmp_path: Path) -> None:
    """Plan AE2. Filing this as an exception concedes a vulnerability we do not have."""
    entry = _one(
        tmp_path,
        FRONTEND,
        row("CVE-9000-0010", "libssl3", "3.3.9-r0", fixed="3.3.7-r0", manager="apk"),
    )
    assert entry.verdict.disposition is Disposition.DISPUTE
    assert "already meets or exceeds" in entry.verdict.reason
    assert "3.3.7-r0" in entry.verdict.reason  # the reviewer can check it without asking us


def test_a_prerelease_only_fix_is_accepted_risk_with_the_prerelease_as_its_reason(
    tmp_path: Path,
) -> None:
    """Plan AE3. There is no released version to move to, so 'deferred' would be a lie —
    deferred means a fix exists and we chose not to take it."""
    entry = _one(
        tmp_path,
        BACKEND,
        row(
            "CVE-9000-0011",
            "python",
            "3.14.6",
            fixed="3.15.0a6",
            manager="binary",
            kind="Application",
        ),
    )
    assert entry.verdict.disposition is Disposition.EXCEPTION
    assert "pre-release" in entry.verdict.reason
    assert "3.15.0a6" in entry.verdict.reason


def test_a_fixable_finding_nobody_claimed_is_deferred_and_loudly_unowned(
    tmp_path: Path,
) -> None:
    """A fixable finding that no rule clears must never quietly become an exception.

    It lands in the third disposition with placeholder owner/date that the CLI treats as a
    blocker — an unowned deferral is indistinguishable from an oversight, so it is made noisy
    rather than tidy.
    """
    entry = _one(
        tmp_path,
        FRONTEND,
        # An ecosystem no frontend rule matches, WITH a released fix available.
        row(
            "CVE-9000-0012",
            "some-lib",
            "1.0.0",
            fixed="1.0.1",
            manager="cargo",
            kind="Application",
        ),
    )
    assert entry.verdict.disposition is Disposition.DEFERRED
    assert entry.verdict.owner == "UNASSIGNED"
    assert entry.verdict.target_date == "UNSET"


def test_debian_no_fix_rows_are_held_never_filed_as_exceptions(tmp_path: Path) -> None:
    """The Debian feed names no fixed version on ANY row, so an empty field there is a property
    of the feed rather than proof the package is unfixable. Filing it as an exception asserts
    something the data does not support."""
    entry = _one(tmp_path, SANDBOX, row("CVE-9000-0013", "libsystemd0", "252.39-1", fixed="-"))
    assert entry.verdict.disposition is Disposition.HELD


def test_the_kept_sandbox_tooling_is_excepted_only_when_it_is_genuinely_unfixable(
    tmp_path: Path,
) -> None:
    """curl and procps stay in the sandbox by decision — but a FIXABLE finding against them is
    still cleared by the base move. Filing a fixable one as an accepted risk while claiming
    zero-fixable is the single claim a reviewer can disprove from their own console."""
    unfixable = _one(
        tmp_path,
        SANDBOX,
        row("CVE-9000-0014", "curl", "7.88.1-10", fixed="-"),
        name="unfixable.xlsx",
    )
    fixable = _one(
        tmp_path,
        SANDBOX,
        row("CVE-9000-0015", "curl", "7.88.1-10", fixed="7.88.1-11"),
        name="fixable.xlsx",
    )
    assert unfixable.verdict.disposition is Disposition.EXCEPTION
    assert fixable.verdict.disposition is Disposition.FIXED


def test_a_mixed_group_takes_the_strictest_verdict_never_the_cleared_one(tmp_path: Path) -> None:
    """Rows sharing an entry can disagree when they sit at different install paths. Reporting
    such a group as Fixed would claim more than was done."""
    path = write_severity_shape(
        tmp_path / "mixed.xlsx",
        asset_ref(SANDBOX),
        [
            # Same CVE + package + version, two paths: one cleared by the esbuild override,
            # one an OS package with no named fix.
            row("CVE-9000-0016", "shared", "1.0", path="/workspace/app/node_modules/esbuild/x"),
            row("CVE-9000-0016", "shared", "1.0", path="/var/lib/dpkg/status"),
        ],
    )
    entries = build_entries(load_export(path))
    assert len(entries) == 1  # one entry, two rows
    assert entries[0].rows_accounted == 2
    assert entries[0].verdict.disposition is not Disposition.FIXED


# ─────────────────────────────────────────────────────────────────────────────
# Granularity and the partition invariant
# ─────────────────────────────────────────────────────────────────────────────


def test_one_cve_across_sibling_packages_and_paths_collapses_per_package(
    tmp_path: Path,
) -> None:
    """Plan AE6, with the real 9-row shape: one CVE, three sibling packages, three paths each.

    The report is per (image, CVE, package, version) and carries the ROW count, so the reviewer
    can reconcile against the row totals their console shows while reading a list of real
    findings rather than nine near-duplicates.
    """
    rows = [
        row("CVE-9000-0020", pkg, "2.41-12", path=p)
        for pkg in ("libc6", "libc-bin", "perl-base")
        for p in ("/var/lib/dpkg/status", "/usr/share/doc/a", "/usr/share/doc/b")
    ]
    path = write_severity_shape(tmp_path / "nine.xlsx", asset_ref(SANDBOX), rows)
    findings = load_export(path)
    entries = build_entries(findings)

    assert len(findings) == 9
    assert len(entries) == 3  # one entry per affected PACKAGE
    assert {e.software_name for e in entries} == {"libc6", "libc-bin", "perl-base"}
    assert all(e.rows_accounted == 3 for e in entries)
    assert sum(e.rows_accounted for e in entries) == 9
    # Every path is preserved on its entry — collapsing must not lose where it was found.
    assert all(len(e.install_paths) == 3 for e in entries)


def test_every_scanner_row_lands_in_exactly_one_entry(tmp_path: Path) -> None:
    """The partition invariant, checked at the row level via provenance."""
    rows = [row(f"CVE-9000-01{i:02d}", f"pkg{i % 4}", "1.0") for i in range(30)]
    path = write_severity_shape(tmp_path / "many.xlsx", asset_ref(SANDBOX), rows)
    findings = load_export(path)

    integrity = check_integrity(findings, build_entries(findings))

    assert integrity.ok, integrity.problems
    assert integrity.total_rows_in == 30
    assert integrity.total_rows_out == 30


def test_integrity_fails_loudly_when_a_row_is_dropped_or_duplicated(tmp_path: Path) -> None:
    """The check must actually be capable of failing — a partition test that cannot go red
    proves nothing. Both directions are exercised: a lost row and a double-counted one."""
    rows = [row(f"CVE-9000-02{i:02d}", "pkg", "1.0") for i in range(5)]
    path = write_severity_shape(tmp_path / "drop.xlsx", asset_ref(SANDBOX), rows)
    findings = load_export(path)
    entries = build_entries(findings)

    dropped = check_integrity(findings, entries[:-1])
    assert not dropped.ok
    assert any("landed in no entry" in p for p in dropped.problems)

    doubled = check_integrity(findings, [*entries, entries[0]])
    assert not doubled.ok
    assert any("more than once" in p for p in doubled.problems)


# ─────────────────────────────────────────────────────────────────────────────
# Reconciliation
# ─────────────────────────────────────────────────────────────────────────────


def _entries(
    tmp_path: Path,
    name: str,
    rows: list[dict[str, str]],
    digest: str,
    *,
    image: str = FRONTEND,
) -> list[Entry]:
    """Entries pinned to a specific asset DIGEST — which the normal factories cannot express.

    Defaults to the frontend image because these are reconciliation tests: they need rows the
    rule table genuinely clears, so that "predicted FIXED" is a real precondition rather than
    an accident. An apk row attributed to the sandbox image matches no rule and lands in the
    fallback, which would make the reconcile assertions vacuous.
    """
    wb_path = tmp_path / name
    wb = Workbook()
    if (default := wb.active) is not None:
        wb.remove(default)
    ws = wb.create_sheet("sheet")
    ws.append(_ASSET_HEADERS)
    for r in rows:
        ws.append(
            [
                r["cve"],
                "d",
                r["severity"],
                "New",
                asset_ref(image, digest),
                r["software"],
                r["version"],
                r["fixed"],
                "2026-08-01",
                r["manager"],
                "False",
            ]
        )
    wb.save(wb_path)
    return build_entries(load_export(wb_path))


def test_a_fix_that_did_not_land_is_reported_as_such_never_as_a_new_exception(
    tmp_path: Path,
) -> None:
    """Laundering a failed remediation into a defended one is the worst outcome available."""
    r = row("CVE-9000-0030", "libexpat1", "2.5.0-1", fixed="2.5.0-2", manager="apk")
    before = _entries(tmp_path, "b.xlsx", [r], DIGEST_OLD)
    after = _entries(tmp_path, "a.xlsx", [r], DIGEST_OLD)
    assert before[0].verdict.disposition is Disposition.FIXED  # predicted cleared

    rec = reconcile(before, after)

    assert len(rec.failed_to_clear) == 1
    assert rec.entries[0].verdict.disposition is Disposition.DEFERRED
    assert "STILL PRESENT" in rec.entries[0].verdict.reason


def test_a_residual_against_an_unshipped_digest_is_superseded_not_a_failed_fix(
    tmp_path: Path,
) -> None:
    """THE CASE THAT DECIDES THE REDUCTION CLAIM.

    A push untags the previous manifest without deleting it. If the scanner enumerates retained
    manifests, the old image keeps reporting its full finding set — and a reconcile that does
    not check the digest reads every one of those rows as 'predicted cleared, still present'
    and reports a successful remediation as a total failure.
    """
    r = row("CVE-9000-0031", "libexpat1", "2.5.0-1", fixed="2.5.0-2", manager="apk")
    before = _entries(tmp_path, "b2.xlsx", [r], DIGEST_OLD)
    after = _entries(tmp_path, "a2.xlsx", [r], DIGEST_OLD)  # scanner still on the OLD manifest

    assert before[0].verdict.disposition is Disposition.FIXED  # predicted cleared

    rec = reconcile(before, after, current_digests={FRONTEND: DIGEST_NEW})

    assert not rec.failed_to_clear, "an unshipped manifest must not read as a failed fix"
    assert len(rec.superseded) == 1
    entry = rec.entries[0]
    assert entry.verdict.disposition is Disposition.SUPERSEDED
    assert entry.replacing_digest == DIGEST_NEW
    assert DIGEST_OLD in entry.verdict.reason  # says WHICH artifact it measured


def test_without_current_digests_the_superseded_check_is_skipped_not_guessed(
    tmp_path: Path,
) -> None:
    """Absence of the digest must not silently change verdicts — it degrades to the honest
    reading (the fix did not land) rather than inventing a more flattering one."""
    r = row("CVE-9000-0032", "libexpat1", "2.5.0-1", fixed="2.5.0-2", manager="apk")
    before = _entries(tmp_path, "b3.xlsx", [r], DIGEST_OLD)
    after = _entries(tmp_path, "a3.xlsx", [r], DIGEST_OLD)

    rec = reconcile(before, after)

    assert not rec.superseded
    assert len(rec.failed_to_clear) == 1


def test_an_anticipated_addition_carries_its_reason_and_an_unanticipated_one_is_flagged(
    tmp_path: Path,
) -> None:
    """Installing git to fix the publish defect ADDS findings during a CVE remediation. Saying
    so plainly is the point; discovering it in the reviewer's console is not."""
    before: list[Entry] = []
    git_row = row("CVE-9000-0033", "git", "2.47.3", manager="deb")
    mystery_row = row("CVE-9000-0034", "who-put-this-here", "9.9", manager="deb")
    wb_path = tmp_path / "after4.xlsx"
    write_asset_shape(wb_path, {BACKEND: (asset_ref(BACKEND), [git_row, mystery_row])})
    after = build_entries(load_export(wb_path))

    rec = reconcile(before, after)

    by_pkg = {e.software_name: e for e in rec.entries}
    assert by_pkg["git"].verdict.disposition is Disposition.EXCEPTION
    assert "knowingly-accepted" in by_pkg["git"].verdict.reason.lower()
    assert len(rec.unanticipated) == 1
    assert "NOT ANTICIPATED" in by_pkg["who-put-this-here"].verdict.reason


# ─────────────────────────────────────────────────────────────────────────────
# Human annotations
# ─────────────────────────────────────────────────────────────────────────────


def test_a_rejected_dispute_falls_back_to_accepted_risk_rather_than_disappearing(
    tmp_path: Path,
) -> None:
    """A dispute round-trip sits on the critical path to approval and can come back negative.
    The rejected row must remain in the report with a stated reason — the failure mode being
    guarded against is a rejected dispute quietly vanishing from the totals."""
    entry = _one(
        tmp_path,
        FRONTEND,
        row("CVE-9000-0040", "libssl3", "3.3.9-r0", fixed="3.3.7-r0", manager="apk"),
    )
    assert entry.verdict.disposition is Disposition.DISPUTE

    unmatched = apply_overrides(
        [entry],
        [
            Override(
                image=FRONTEND,
                cve_id="CVE-9000-0040",
                software_name="libssl3",
                disposition=Disposition.DISPUTE_REJECTED,
                reason="Scan owner rejected the dispute on 2026-09-01; accepted as risk.",
            )
        ],
    )

    assert unmatched == []
    # Re-read through a fresh local: `apply_overrides` REPLACES `entry.verdict`, which the type
    # checker cannot see, so asserting on the same attribute expression twice reads as a
    # contradiction rather than as a before/after.
    annotated = entry.verdict
    assert annotated.disposition is Disposition.DISPUTE_REJECTED
    assert "rejected" in annotated.reason.lower()


def test_an_override_that_matches_nothing_is_reported_not_ignored(tmp_path: Path) -> None:
    """The realistic mistake is a typo in a CVE id, and the symptom of ignoring it is a report
    silently missing the exact row somebody went to the trouble of annotating."""
    entry = _one(tmp_path, FRONTEND, row("CVE-9000-0041", "libssl3", "3.3.3-r0", manager="apk"))

    unmatched = apply_overrides(
        [entry],
        [Override(image=FRONTEND, cve_id="CVE-9000-9999", software_name="libssl3")],
    )

    assert unmatched == [f"{FRONTEND}/CVE-9000-9999/libssl3"]


def test_an_override_supplies_the_owner_and_date_a_deferral_requires(tmp_path: Path) -> None:
    entry = _one(
        tmp_path,
        FRONTEND,
        row("CVE-9000-0042", "some-lib", "1.0.0", fixed="1.0.1", manager="cargo"),
    )
    assert entry.verdict.owner == "UNASSIGNED"

    apply_overrides(
        [entry],
        [
            Override(
                image=FRONTEND,
                cve_id="CVE-9000-0042",
                software_name="some-lib",
                owner="Platform team",
                target_date="2026-10-31",
                vendor_status="Under investigation",
                tracker_url="https://example.invalid/tracker/1",
            )
        ],
    )

    assert entry.verdict.disposition is Disposition.DEFERRED  # unchanged — only annotated
    assert entry.verdict.owner == "Platform team"
    assert entry.verdict.target_date == "2026-10-31"
    assert entry.tracker_url == "https://example.invalid/tracker/1"


def test_an_unknown_disposition_in_the_annotations_file_fails_loudly(tmp_path: Path) -> None:
    path = tmp_path / "overrides.json"
    path.write_text(
        json.dumps(
            {
                "entries": [
                    {
                        "image": FRONTEND,
                        "cve_id": "CVE-9000-0043",
                        "software_name": "libssl3",
                        "disposition": "not-affected",  # the one word this report may never use
                    }
                ]
            }
        ),
        encoding="utf-8",
    )
    with pytest.raises(ValueError, match="unknown disposition"):
        load_overrides(path)


def test_a_missing_required_field_in_the_annotations_file_fails_loudly(tmp_path: Path) -> None:
    path = tmp_path / "bad.json"
    path.write_text(json.dumps({"entries": [{"image": FRONTEND, "cve_id": "CVE-9000-0044"}]}))
    with pytest.raises(KeyError):
        load_overrides(path)


# ─────────────────────────────────────────────────────────────────────────────
# The workbook
# ─────────────────────────────────────────────────────────────────────────────

_SUMIFS = re.compile(r"SUMIFS\(([^()]*)\)")
_REF = re.compile(r"'([^']+)'!\$([A-Z]+):\$[A-Z]+")


def _eval_sumifs(wb: Workbook, formula: str) -> int:
    """Evaluate the SUMIFS terms this generator writes, so 'derived' can actually be proven.

    openpyxl stores formulas, it does not compute them — so a test that only asserted the cell
    holds a string starting with `=` would prove the cell is A formula, not that it is the RIGHT
    one. This computes the formula against the written sheets, which is what makes
    'Summary moves when a sheet moves' a real assertion rather than a shape check.
    """
    total = 0
    for args_text in _SUMIFS.findall(formula):
        args = [a.strip() for a in args_text.split(",")]
        sum_match = _REF.match(args[0])
        assert sum_match, args[0]
        sum_sheet, sum_col = sum_match.group(1), sum_match.group(2)
        criteria: list[tuple[str, str]] = []
        for i in range(1, len(args) - 1, 2):
            ref = _REF.match(args[i])
            assert ref, args[i]
            criteria.append((ref.group(2), args[i + 1].strip('"')))
        ws = wb[sum_sheet]
        for r in range(2, (ws.max_row or 1) + 1):
            if all(str(ws[f"{col}{r}"].value or "") == want for col, want in criteria):
                total += int(ws[f"{sum_col}{r}"].value or 0)
    return total


def _build(tmp_path: Path, entries: list[Entry], before_rows: dict[str, int]) -> Path:
    wb = build_workbook(
        entries,
        before_rows=before_rows,
        after_rows=dict.fromkeys(before_rows),
        digests={SANDBOX: DIGEST_OLD},
        generated_at="2026-08-13 00:00 UTC",
    )
    out = tmp_path / "register.xlsx"
    wb.save(out)
    return out


def test_the_workbook_partitions_every_row_across_the_sheets(tmp_path: Path) -> None:
    """Every row from the original list appears in exactly ONE sheet, and the accounted counts
    sum back to the scanner's own total. This is the report's self-check."""
    rows = [
        row("CVE-9000-0050", "libexpat1", "2.5.0-1", fixed="2.5.0-2"),  # fixed by the base move
        row("CVE-9000-0051", "libsystemd0", "252.39-1", fixed="-"),  # held (debian, no fix)
        row("CVE-9000-0052", "curl", "7.88.1-10", fixed="-"),  # excepted (kept tooling)
        row(
            "CVE-9000-0053", "stdlib", "go1.20.7", manager="go-module", path="/usr/local/bin/caddy"
        ),
    ]
    src = write_severity_shape(tmp_path / "src.xlsx", asset_ref(SANDBOX), rows)
    findings = load_export(src)
    entries = build_entries(findings)
    out = _build(tmp_path, entries, {SANDBOX: len(findings)})

    wb = openpyxl.load_workbook(out)
    content = ["Fixed", "Exceptions", "Deferred", "Disputes", "Out of scope", "Held"]
    assert wb.sheetnames[0] == "Summary"  # the sheet most readers will open
    assert set(content).issubset(set(wb.sheetnames))

    seen: list[tuple[str, str]] = []
    accounted = 0
    for name in content:
        ws = wb[name]
        for r in range(2, (ws.max_row or 1) + 1):
            if ws[f"A{r}"].value is None:
                continue
            seen.append((str(ws[f"B{r}"].value), str(ws[f"D{r}"].value)))
            accounted += int(ws[f"J{r}"].value or 0)

    assert len(seen) == len(set(seen)), "a finding appears in more than one sheet"
    assert accounted == len(findings), "the sheets do not sum back to the scanner's row total"


def test_summary_is_derived_from_the_sheets_and_moves_when_they_do(tmp_path: Path) -> None:
    """`Summary` must be a formula over the other sheets, never typed.

    Proven in two steps, because either alone is weak: the cell must BE a formula (a typed
    literal fails here), and evaluating that formula against a MUTATED sheet must produce the
    mutated number (a formula pointing at the wrong range fails here).
    """
    rows = [
        row(f"CVE-9000-006{i}", "libexpat1", f"2.5.0-{i}", fixed=f"2.5.0-{i + 1}")
        for i in range(4)
    ]
    src = write_severity_shape(tmp_path / "s.xlsx", asset_ref(SANDBOX), rows)
    findings = load_export(src)
    out = _build(tmp_path, build_entries(findings), {SANDBOX: len(findings)})

    wb = openpyxl.load_workbook(out)
    summary = wb["Summary"]
    header_row = next(
        r for r in range(1, 30) if str(summary[f"A{r}"].value or "").strip() == "Image"
    )
    fixed_cell = summary.cell(row=header_row + 1, column=6)

    # (1) It is a formula, not a number somebody typed.
    assert isinstance(fixed_cell.value, str) and fixed_cell.value.startswith("=SUMIFS("), (
        f"Summary!F{header_row + 1} is typed, not derived: {fixed_cell.value!r}"
    )

    # (2) Evaluating it reproduces the sheet, and TRACKS the sheet when the sheet changes.
    before = _eval_sumifs(wb, fixed_cell.value)
    assert before == len(findings)

    wb["Fixed"]["J2"] = int(wb["Fixed"]["J2"].value or 0) + 7  # hand-edit one row
    after = _eval_sumifs(wb, fixed_cell.value)
    assert after == before + 7, "Summary did not move with the sheet it claims to summarise"


def test_held_rows_never_leak_into_a_residual_sheet_they_do_not_belong_to(
    tmp_path: Path,
) -> None:
    """Held rows are counted and visible, but they must not be laundered into Exceptions —
    that would assert a defence we have not established."""
    rows = [row(f"CVE-9000-007{i}", f"libpkg{i}", "1.0", fixed="-") for i in range(3)]
    src = write_severity_shape(tmp_path / "h.xlsx", asset_ref(SANDBOX), rows)
    findings = load_export(src)
    out = _build(tmp_path, build_entries(findings), {SANDBOX: len(findings)})

    wb = openpyxl.load_workbook(out)
    held = wb["Held"]
    assert (held.max_row or 1) - 1 == 3
    for name in ("Exceptions", "Deferred", "Disputes", "Out of scope", "Fixed"):
        ws = wb[name]
        values = [ws[f"K{r}"].value for r in range(2, (ws.max_row or 1) + 1)]
        assert str(Disposition.HELD) not in [str(v) for v in values]


def _summary_of(wb: Workbook, out: Path) -> tuple[Worksheet, int]:
    """Save, reload, and return the Summary sheet with the index of its first image row."""
    wb.save(out)
    loaded = openpyxl.load_workbook(out)["Summary"]
    header_row = next(
        r for r in range(1, 30) if str(loaded[f"A{r}"].value or "").strip() == "Image"
    )
    return loaded, header_row + 1


def test_an_image_that_was_never_scanned_carries_an_after_only(tmp_path: Path) -> None:
    """The deployed-app image has no 'before' — its Summary row must SAY so rather than show a
    blank that reads as zero findings.

    `never_scanned` is what makes this THAT case. Without it, before-data-and-no-after is the
    pre-rescan case below, which is a different claim entirely.
    """
    rows = [row("CVE-9000-0080", "libexpat1", "2.5.0-1", fixed="2.5.0-2")]
    src = write_severity_shape(tmp_path / "n.xlsx", asset_ref(SANDBOX), rows)
    entries = build_entries(load_export(src))
    wb = build_workbook(
        entries,
        before_rows={},
        after_rows={SANDBOX: None},
        digests={},
        generated_at="2026-08-13 00:00 UTC",
        never_scanned=[SANDBOX],
    )
    loaded, first = _summary_of(wb, tmp_path / "never.xlsx")
    assert loaded.cell(row=first, column=3).value == "not scanned"
    assert loaded.cell(row=first, column=4).value == "never scanned"
    assert loaded.cell(row=first, column=2).value == "not scanned"


def test_before_data_with_no_after_reads_as_pending_not_as_never_scanned(
    tmp_path: Path,
) -> None:
    """THE PRE-RESCAN WORKBOOK. `map` builds this before BIAL rescans, because the images are
    pushed from BIAL's own Windows VM and the delivery team cannot trigger the rescan.

    Both cases arrive as `after is None`, and conflating them puts the wrong claim in front of a
    client: "never scanned" about an image we measured in full says we never looked at it. Only
    `never_scanned` distinguishes them.
    """
    rows = [row("CVE-9000-0081", "libexpat1", "2.5.0-1", fixed="2.5.0-2")]
    src = write_severity_shape(tmp_path / "p.xlsx", asset_ref(SANDBOX), rows)
    entries = build_entries(load_export(src))
    wb = build_workbook(
        entries,
        before_rows={SANDBOX: 806},
        after_rows={SANDBOX: None},
        digests={SANDBOX: DIGEST_OLD},
        generated_at="2026-08-16 00:00 UTC",
        before_files=["vibe-coding_sheet.xlsx"],
        after_files=(),
    )
    loaded, first = _summary_of(wb, tmp_path / "pending.xlsx")

    assert loaded.cell(row=first, column=3).value == 806, "the before count is real and measured"
    assert loaded.cell(row=first, column=4).value == "pending rescan"
    assert loaded.cell(row=first, column=5).value == "pending rescan"
    assert "NOT YET RECEIVED" in str(loaded["A3"].value), (
        "the header must say the post-remediation export has not arrived"
    )


def test_no_finding_is_ever_labelled_not_affected(tmp_path: Path) -> None:
    """A guard on the vocabulary itself. 'Not affected' is the one claim this report may never
    make about a fixable finding, and it is the easiest one to reach for by accident."""
    assert "not-affected" not in {d.value for d in Disposition}
    rows = [
        row("CVE-9000-0090", "a", "1.0", fixed="2.0"),
        row("CVE-9000-0091", "b", "1.0", fixed="-"),
        row("CVE-9000-0092", "c", "1.0", fixed="9.9.9a1"),
    ]
    src = write_severity_shape(tmp_path / "v.xlsx", asset_ref(SANDBOX), rows)
    for entry in build_entries(load_export(src)):
        assert "not affected" not in entry.verdict.reason.lower()
        assert "not-affected" not in str(entry.verdict.disposition)


# ─────────────────────────────────────────────────────────────────────────────
# The CLI — where the exit-code contract lives
# ─────────────────────────────────────────────────────────────────────────────
#
# WHY THIS SECTION EXISTS. Every test above calls an internal function directly, so `main()` —
# argparse wiring, the `map`/`register` split, `--current-digest` parsing, and the ready/not-ready
# exit code an operator reads to decide whether a report is safe to hand to the client — had no
# coverage at all. Four defects lived in exactly that gap, and each one made the tool assert
# something friendlier than the truth:
#
#   * `register` with no `--after` re-dispositioned every row to `fixed` and exited 0.
#   * A `--current-digest` matching no manifest in the export filed the whole residual set as
#     superseded and exited 0.
#   * The HELD gate was snapshotted before reconcile, so it blocked a cleared run forever.
#   * The Summary's partition check read MISMATCH whenever an anticipated addition existed.
#
# The shared property: a report that overstates remediation must not be reachable with a zero
# exit code. That is what these pin.


def _register_inputs(tmp_path: Path) -> tuple[Path, Path]:
    """A before-export whose row the rule table clears, and an after-export where it is gone."""
    r = row("CVE-9000-0300", "libexpat1", "2.5.0-1", fixed="2.5.0-2", manager="apk")
    before = write_asset_shape(tmp_path / "before.xlsx", {FRONTEND: (asset_ref(FRONTEND), [r])})
    after = write_asset_shape(
        tmp_path / "after.xlsx",
        {FRONTEND: (asset_ref(FRONTEND, DIGEST_NEW), [row("CVE-9000-0301", "other", "1.0")])},
    )
    return before, after


def test_register_refuses_to_run_without_a_post_remediation_export(tmp_path: Path) -> None:
    """The worst reachable output: a report claiming 100% remediation against data never seen.

    `reconcile` reads "absent from the after-export" as "cleared", so an empty after-set turned
    every exception, deferral and held row into `fixed` with the reason "cleared as a side effect
    of this pass" — and exited 0 while the Summary said the post-remediation export had not
    arrived. `--after` is required, and the operator is pointed at `map`.
    """
    before, _ = _register_inputs(tmp_path)
    with pytest.raises(SystemExit) as exit_info:
        main(["register", "--before", str(before), "--out", str(tmp_path / "out.xlsx")])
    assert exit_info.value.code == 2  # argparse's own "required argument" exit


def test_register_refuses_an_after_export_that_parses_to_nothing(tmp_path: Path) -> None:
    """Passing `--after` is not the same as it containing findings. An empty workbook takes the
    same path as omitting the flag, so it gets the same refusal rather than the flattering read."""
    before, _ = _register_inputs(tmp_path)
    empty = write_asset_shape(tmp_path / "empty.xlsx", {FRONTEND: (asset_ref(FRONTEND), [])})

    assert (
        main(
            [
                "register",
                "--before",
                str(before),
                "--after",
                str(empty),
                "--out",
                str(tmp_path / "o"),
            ]
        )
        == 2
    )


def _two_image_before(tmp_path: Path) -> Path:
    """A before-export covering TWO images, so an after-export can plausibly omit one."""
    return write_asset_shape(
        tmp_path / "before2.xlsx",
        {
            FRONTEND: (
                asset_ref(FRONTEND),
                [row("CVE-9000-0320", "libexpat1", "2.5.0-1", fixed="2.5.0-2", manager="apk")],
            ),
            BACKEND: (
                asset_ref(BACKEND),
                [row("CVE-9000-0321", "libssl3", "3.3.3-r0", fixed="3.3.7-r0", manager="apk")],
            ),
        },
    )


def test_an_image_missing_from_the_after_export_is_refused_not_scored_as_zero(
    tmp_path: Path,
) -> None:
    """THE STAGGERED-RESCAN PATH. A missing after-count used to be filled with a typed 0, and zero
    residual rows renders as a clean sweep: full reduction for that image, its HELD rows rewritten
    to `fixed`, and the ship-gate satisfied because the rows it guards stopped existing.

    BIAL returning two images this week and the third next week is the ORDINARY case, so this
    fired on a normal Tuesday and claimed an image was cleared that nobody had looked at.
    """
    before = _two_image_before(tmp_path)
    # The after-export covers the FRONTEND only — the BACKEND was simply not rescanned yet. The
    # FRONTEND must carry a REAL residual row: an after-export that parses to nothing trips the
    # empty-export guard above instead, and this test would pass without exercising its own guard.
    after = write_asset_shape(
        tmp_path / "after1.xlsx",
        {
            FRONTEND: (
                asset_ref(FRONTEND, DIGEST_NEW),
                [row("CVE-9000-0322", "zlib1g", "1.3-r0", fixed="1.4-r0", manager="apk")],
            )
        },
    )

    assert (
        main(
            [
                "register",
                "--before",
                str(before),
                "--after",
                str(after),
                "--out",
                str(tmp_path / "partial.xlsx"),
            ]
        )
        == 2  # fmt: skip
    ), "an unrescanned image must not be scored as fully remediated"


def test_the_missing_image_is_allowed_once_the_operator_declares_it(tmp_path: Path) -> None:
    """The guard must not wall off the honest path. `--never-scanned` already encodes 'no data
    here' as None — rendered as text, never as a numeric zero — so declaring the gap proceeds.
    Without this, the guard above would make a genuinely staggered rescan unreportable."""
    before = _two_image_before(tmp_path)
    after = write_asset_shape(
        tmp_path / "after2.xlsx",
        {
            FRONTEND: (
                asset_ref(FRONTEND, DIGEST_NEW),
                [row("CVE-9000-0323", "zlib1g", "1.3-r0", fixed="1.4-r0", manager="apk")],
            )
        },
    )

    code = main([
        "register", "--before", str(before), "--after", str(after),
        "--out", str(tmp_path / "declared.xlsx"), "--never-scanned", BACKEND,
    ])  # fmt: skip
    assert code != 2, "declaring the gap is the sanctioned way to report a staggered rescan"


def test_a_current_digest_the_scan_never_saw_is_refused_not_trusted(tmp_path: Path) -> None:
    """THE LAUNDERING PATH. A digest matching no manifest in the export marks every residual row
    `superseded-artifact` — dropping it out of the reduction — so a total remediation failure
    reads as a clean report with a zero exit. The realistic trigger is not a typo: it is pasting
    an INDEX digest where the scanner reports the per-architecture child.
    """
    r = row("CVE-9000-0310", "libssl3", "3.3.3-r0", fixed="3.3.7-r0", manager="apk")
    before = write_asset_shape(tmp_path / "b.xlsx", {FRONTEND: (asset_ref(FRONTEND), [r])})
    # Every row survived — the remediation did nothing.
    after = write_asset_shape(tmp_path / "a.xlsx", {FRONTEND: (asset_ref(FRONTEND), [r])})
    out = tmp_path / "r.xlsx"

    honest = main(["register", "--before", str(before), "--after", str(after), "--out", str(out)])
    assert honest == 1, "a remediation that cleared nothing must not report success"

    assert main([
        "register", "--before", str(before), "--after", str(after), "--out", str(out),
        "--current-digest", f"{FRONTEND}=sha256:{'f' * 64}",
    ]) == 2  # fmt: skip


def test_a_current_digest_the_scan_did_see_still_supersedes(tmp_path: Path) -> None:
    """The guard must not break the case `--current-digest` exists for: a scanner enumerating
    retained manifests, where the residual rows genuinely describe an artifact we stopped
    shipping. The digest is present in the export, so it is trusted."""
    r = row("CVE-9000-0311", "libssl3", "3.3.3-r0", fixed="3.3.7-r0", manager="apk")
    before = write_asset_shape(tmp_path / "b2.xlsx", {FRONTEND: (asset_ref(FRONTEND), [r])})
    after = write_asset_shape(
        tmp_path / "a2.xlsx",
        {
            "old": (asset_ref(FRONTEND, DIGEST_OLD), [r]),
            "new": (asset_ref(FRONTEND, DIGEST_NEW), [row("CVE-9000-0312", "unrelated", "1.0")]),
        },
    )
    out = tmp_path / "r2.xlsx"

    main([
        "register", "--before", str(before), "--after", str(after), "--out", str(out),
        "--current-digest", f"{FRONTEND}={DIGEST_NEW}",
    ])  # fmt: skip

    sheet = openpyxl.load_workbook(out)["Out of scope"]
    dispositions = [sheet[f"K{n}"].value for n in range(2, (sheet.max_row or 1) + 1)]
    assert str(Disposition.SUPERSEDED) in [str(d) for d in dispositions]


def test_the_held_gate_reads_the_workbook_that_was_written_not_the_coverage_map(
    tmp_path: Path,
) -> None:
    """`integrity` is computed before `reconcile` mutates the same Entry objects, so gating on
    its HELD list described a report that no longer existed: once any row had ever been HELD, no
    run could exit 0 again — even with an empty Held sheet. The gate now reads `rec.entries`.
    """
    # A Debian no-fix row is HELD in the coverage map; it is absent from the rescan, so it clears.
    # The kept-tooling row survives as a residual EXCEPTION — present in both exports, so it is
    # neither an addition nor an unowned deferral, leaving HELD as the only thing under test.
    held_row = row("CVE-9000-0320", "libsystemd0", "252.39-1", fixed="-")
    survives = row("CVE-9000-0321", "curl", "7.88.1-10", fixed="-")
    before = write_severity_shape(tmp_path / "b3.xlsx", asset_ref(SANDBOX), [held_row, survives])
    after = write_severity_shape(tmp_path / "a3.xlsx", asset_ref(SANDBOX), [survives])
    out = tmp_path / "r3.xlsx"

    by_cve = {e.cve_id: e for e in build_entries(load_export(before))}
    assert by_cve["CVE-9000-0320"].verdict.disposition is Disposition.HELD
    assert by_cve["CVE-9000-0321"].verdict.disposition is Disposition.EXCEPTION

    code = main(["register", "--before", str(before), "--after", str(after), "--out", str(out)])

    assert (openpyxl.load_workbook(out)["Held"].max_row or 1) - 1 == 0, "nothing is still held"
    assert code == 0, "a run whose Held sheet emptied must be shippable"


def test_a_row_still_held_after_the_rescan_blocks_the_run(tmp_path: Path) -> None:
    """The other direction — the gate has to stay capable of firing."""
    held_row = row("CVE-9000-0330", "libsystemd0", "252.39-1", fixed="-")
    before = write_severity_shape(tmp_path / "b4.xlsx", asset_ref(SANDBOX), [held_row])
    after = write_severity_shape(tmp_path / "a4.xlsx", asset_ref(SANDBOX), [held_row])
    out = tmp_path / "r4.xlsx"

    assert (
        main(["register", "--before", str(before), "--after", str(after), "--out", str(out)]) == 1
    )
    assert (openpyxl.load_workbook(out)["Held"].max_row or 1) - 1 == 1


def test_the_partition_check_stays_ok_when_an_anticipated_addition_appears(
    tmp_path: Path,
) -> None:
    """THE REPORT'S HEADLINE SELF-CHECK, on the one case the tool is built to anticipate.

    Installing git to fix the publish defect ADDS findings. Those rows are written to the content
    sheets but have no counterpart in the before-export, so comparing `Sheets total` against
    `Before rows` alone rendered MISMATCH for a correct report — on the backend row and on the
    grand TOTAL — while the CLI printed `partition : OK`. A self-check that cries wolf on its own
    designed-for case teaches the reviewer to ignore the column.
    """
    cleared = row("CVE-9000-0340", "libc6", "2.41-12", fixed="-", manager="deb")
    before = write_asset_shape(tmp_path / "b5.xlsx", {BACKEND: (asset_ref(BACKEND), [cleared])})
    # The row cleared; `git` arrived. One before-row, one addition, two sheet rows.
    after = write_asset_shape(
        tmp_path / "a5.xlsx",
        {BACKEND: (asset_ref(BACKEND), [row("CVE-9000-0341", "git", "2.47.3", manager="deb")])},
    )
    out = tmp_path / "r5.xlsx"
    main(["register", "--before", str(before), "--after", str(after), "--out", str(out)])

    wb = openpyxl.load_workbook(out)
    summary = wb["Summary"]
    header = next(n for n in range(1, 30) if str(summary[f"A{n}"].value or "").strip() == "Image")
    data_row = header + 1
    # openpyxl types a cell value as a wide union; these two are written as ints by
    # `_write_summary`, and the arithmetic below is the whole point of the test.
    before_rows = int(str(summary.cell(row=data_row, column=3).value))
    additions = int(str(summary.cell(row=data_row, column=12).value))
    sheets_total = _eval_sumifs(wb, str(summary.cell(row=data_row, column=13).value))

    assert before_rows == 1
    assert additions == 1, "the git row must be counted as an addition, not lost"
    assert sheets_total == 2, "both the cleared row and the addition are written to sheets"
    # The formula the reviewer reads is `=IF(M=C+L,...)`; evaluate its arms.
    assert sheets_total == before_rows + additions, "Partition check would render MISMATCH"
    assert "C" in str(summary.cell(row=data_row, column=14).value)


def test_the_map_pass_writes_a_row_level_coverage_map(tmp_path: Path) -> None:
    """U9's deliverable. Row level, not entry level, so the reviewer can check it line by line
    against their own console — and so every row demonstrably carries exactly one disposition."""
    rows = [row(f"CVE-9000-035{i}", f"pkg{i}", "1.0", fixed="-") for i in range(3)]
    before = write_severity_shape(tmp_path / "b6.xlsx", asset_ref(SANDBOX), rows)
    out_dir = tmp_path / "coverage"

    code = main(["map", "--before", str(before), "--out-dir", str(out_dir)])

    assert code == 0
    written = (out_dir / "coverage-map.csv").read_text(encoding="utf-8").strip().splitlines()
    assert len(written) == 4, "one header plus one line per scanner row"
    assert all("," in line for line in written)
    summary = json.loads((out_dir / "coverage-summary.json").read_text(encoding="utf-8"))
    assert summary["totals"]["scanner_rows"] == 3
    assert summary["images"][SANDBOX]["rows"] == 3
    assert summary["images"][SANDBOX]["digest"] == DIGEST_OLD


def test_the_map_pass_also_writes_a_submittable_workbook(tmp_path: Path) -> None:
    """The CSV is for us; the workbook is what the client can actually receive.

    This has to exist BEFORE the rescan because of how the engagement is sequenced: the images
    are built and pushed from BIAL's own Windows VM, so the delivery team cannot trigger the
    rescan that `register` needs. Making the only client-facing output depend on that rescan left
    the person who has to submit something with nothing to submit.
    """
    rows = [row(f"CVE-9000-036{i}", f"pkg{i}", "1.0", fixed="-") for i in range(3)]
    before = write_severity_shape(tmp_path / "b7.xlsx", asset_ref(SANDBOX), rows)
    out_dir = tmp_path / "coverage"

    assert main(["map", "--before", str(before), "--out-dir", str(out_dir)]) == 0

    book = out_dir / "coverage-register.xlsx"
    assert book.exists(), "the map pass must leave a workbook, not only a CSV"
    loaded = openpyxl.load_workbook(book)
    assert "Summary" in loaded.sheetnames
    summary = loaded["Summary"]
    assert "NOT YET RECEIVED" in str(summary["A3"].value), (
        "a pre-rescan workbook must say so on its face, not imply a measured reduction"
    )
    header_row = next(
        r for r in range(1, 30) if str(summary[f"A{r}"].value or "").strip() == "Image"
    )
    assert summary.cell(row=header_row + 1, column=4).value == "pending rescan"


def test_the_coverage_map_gives_every_row_exactly_one_disposition(tmp_path: Path) -> None:
    """The checkable definition of done: at handover, no row may be unmapped."""
    rows = [
        row("CVE-9000-0360", "libexpat1", "2.5.0-1", fixed="2.5.0-2"),
        row("CVE-9000-0361", "curl", "7.88.1-10", fixed="-"),
        row(
            "CVE-9000-0362", "stdlib", "go1.20.7", manager="go-module", path="/usr/local/bin/caddy"
        ),
    ]
    before = write_severity_shape(tmp_path / "b7.xlsx", asset_ref(SANDBOX), rows)
    findings = load_export(before)
    out_dir = tmp_path / "cov"

    write_coverage_map(findings, build_entries(findings), out_dir, generated_at="2026-08-13 00:00")

    lines = (out_dir / "coverage-map.csv").read_text(encoding="utf-8").strip().splitlines()
    header = lines[0].split(",")
    disposition_at = header.index("disposition")
    dispositions = [line.split(",")[disposition_at] for line in lines[1:]]
    assert len(dispositions) == 3
    assert all(d in {d2.value for d2 in Disposition} for d in dispositions), dispositions


def test_disposition_for_is_total_every_row_gets_exactly_one(tmp_path: Path) -> None:
    """No row may fall through the rule table unmapped — 'every row carries a disposition' is
    the definition of done the whole engagement is measured against."""
    shapes = [
        row("CVE-9000-0100", "weird", "1.0", manager="conda", kind="Application", path="/opt/x"),
        row("CVE-9000-0101", "", "", fixed="", manager="", kind="", path=""),
        row("CVE-9000-0102", "nginx", "1.27.5-r1", manager="apk"),
    ]
    src = write_severity_shape(tmp_path / "t.xlsx", asset_ref(SANDBOX), shapes)
    for finding in load_export(src):
        verdict = disposition_for(finding)
        assert isinstance(verdict.disposition, Disposition)
        assert verdict.reason
        assert verdict.rule
