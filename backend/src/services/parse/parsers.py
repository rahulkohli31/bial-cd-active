"""File parsers for the app parse endpoint (R26) — spreadsheet → structured rows,
docx → text. DISTINCT from `services/extract/office.py` (which produces Markdown for
the chat model): this produces the `{kind:'spreadsheet', columns, rows, ...}` shape
the deployed app's `BIALData.parseFile` consumes.

The four bounds (untrusted-file-parsing learning): (1) the decoded-size cap is enforced
by the caller before parsing (the attachments upload limits — the old per-app parse HTTP
endpoint was retired with the open-sandbox pivot, but this parse SERVICE stays, driven by
`attachments/router.py`); (2) the zip-bomb guard + (structural gate) run here BEFORE any
inflate — reusing Plan A's shared `zip_safety` + `office` structural validators;
(3) a row/col range-clamp is applied BEFORE iterating; (4) the whole dispatch runs
inside the killable process governor (`governor.py`). Errors are the shared
`FileParseError` (carries status + code, e.g. 413/`FILE_TOO_LARGE`,
415/`UNSUPPORTED_TYPE`).
"""

from __future__ import annotations

import csv
import datetime
import io
from typing import Any

import openpyxl

from src.services.extract.office import (
    EXCEL_MEDIA_TYPE,
    WORD_MEDIA_TYPE,
    OfficeExtractError,
    assert_office_structure,
    extract_office,
)
from src.services.extract.zip_safety import FileParseError, assert_zip_not_bomb, looks_like_zip

# Bound (3): the clamp box, applied before iterating (Express MAX_PARSE_ROWS/COLS).
MAX_PARSE_ROWS = 50_000
MAX_PARSE_COLS = 512

# Content-type ↔ parse-kind (Express `parseKindFor`). Extension wins for CSV so a
# `.csv` mislabelled as ms-excel/plain stays CSV.
_XLSX_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_XLS_TYPE = "application/vnd.ms-excel"


def parse_kind_for(content_type: str, filename: str) -> str | None:
    """Resolve the parse kind, or None (→ 415). Extension wins for CSV."""
    name = (filename or "").lower()
    if name.endswith(".csv") or content_type == "text/csv":
        return "csv"
    if name.endswith(".xlsx") or content_type == _XLSX_TYPE:
        return "xlsx"
    if name.endswith(".xls") or content_type == _XLS_TYPE:
        return "xls"
    if name.endswith(".docx") or content_type == WORD_MEDIA_TYPE:
        return "word"
    return None


def _coerce(value: Any) -> Any:
    """Cell value → JSON-safe: dates to ISO strings; numbers/bools/strings preserved
    (numbers stay numeric for charts; text-formatted identifiers stay text)."""
    if value is None:
        return None
    if isinstance(value, (datetime.datetime, datetime.date, datetime.time)):
        return value.isoformat()
    return value


def _dedupe(name: str, seen: dict[str, int]) -> str:
    if name in seen:
        seen[name] += 1
        return f"{name} ({seen[name]})"
    seen[name] = 0
    return name


def _empty_sheet(sheets: list[str], sheet: str) -> dict[str, Any]:
    return {
        "kind": "spreadsheet",
        "sheets": sheets,
        "sheet": sheet,
        "columns": [],
        "rows": [],
        "rowCount": 0,
        "totalRows": 0,
        "truncated": False,
        "truncationNote": "",
    }


def _truncation_note(total_rows: int, shown: int, total_cols: int, shown_cols: int) -> str:
    parts: list[str] = []
    if total_rows > shown:
        parts.append(f"Showing the first {shown} of {total_rows} rows.")
    if total_cols > shown_cols:
        parts.append(f"Showing the first {shown_cols} of {total_cols} columns.")
    return " ".join(parts)[:800]


def _sheet_to_rows(ws: Any, sheets: list[str]) -> dict[str, Any]:
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    if max_row == 0 or max_col == 0:
        return _empty_sheet(sheets, ws.title)
    # Bound (3): clamp columns BEFORE building the anchor map / iterating.
    end_col = min(max_col, MAX_PARSE_COLS)

    # Merged-cell anchor map (within the clamp) + full-width horizontal banner rows.
    anchors: dict[tuple[int, int], tuple[int, int]] = {}
    banner_rows: set[int] = set()
    for rng in ws.merged_cells.ranges:
        if rng.min_row == rng.max_row and rng.min_col <= 1 and rng.max_col >= end_col:
            banner_rows.add(rng.min_row)
        if rng.min_row > MAX_PARSE_ROWS + 1:
            continue
        for r in range(rng.min_row, min(rng.max_row, MAX_PARSE_ROWS + 1) + 1):
            for c in range(rng.min_col, min(rng.max_col, end_col) + 1):
                anchors[(r, c)] = (rng.min_row, rng.min_col)

    def value_at(r: int, c: int) -> Any:
        anchor_r, anchor_c = anchors.get((r, c), (r, c))
        return _coerce(ws.cell(row=anchor_r, column=anchor_c).value)

    # Skip leading full-width banner rows to find the real header (Express effectiveHeaderRow).
    header_row = 1
    while header_row in banner_rows and header_row < max_row:
        header_row += 1

    seen: dict[str, int] = {}
    columns: list[str] = []
    for c in range(1, end_col + 1):
        raw = value_at(header_row, c)
        base = str(raw).strip() if raw not in (None, "") else ""
        columns.append(_dedupe(base or f"Column {c}", seen))

    total_rows = max(0, max_row - header_row)
    shown = min(total_rows, MAX_PARSE_ROWS)
    rows: list[dict[str, Any]] = []
    for r in range(header_row + 1, header_row + 1 + shown):
        rows.append({columns[c - 1]: value_at(r, c) for c in range(1, end_col + 1)})

    truncated = total_rows > shown or max_col > end_col
    return {
        "kind": "spreadsheet",
        "sheets": sheets,
        "sheet": ws.title,
        "columns": columns,
        "rows": rows,
        "rowCount": len(rows),
        "totalRows": total_rows,
        "truncated": truncated,
        "truncationNote": _truncation_note(total_rows, shown, max_col, end_col)
        if truncated
        else "",
    }


def parse_spreadsheet(buffer: bytes, sheet_name: str | None) -> dict[str, Any]:
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(buffer), data_only=True)
    except Exception as exc:
        raise FileParseError(f"Could not read the spreadsheet: {exc}") from exc
    sheets = workbook.sheetnames
    if not sheets:
        raise FileParseError("The spreadsheet has no worksheets.")
    target = sheet_name if sheet_name else sheets[0]
    if target not in sheets:
        raise FileParseError("The requested sheet was not found.", code="SHEET_NOT_FOUND")
    result = _sheet_to_rows(workbook[target], sheets)
    workbook.close()
    return result


def parse_csv(buffer: bytes, _sheet_name: str | None) -> dict[str, Any]:
    text = buffer.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    collected: list[list[str]] = []
    for index, row in enumerate(reader):
        if index > MAX_PARSE_ROWS:  # header + MAX_PARSE_ROWS data rows
            break
        collected.append(row)
    if not collected:
        return _empty_sheet(["Sheet1"], "Sheet1")

    header = collected[0]
    end_col = min(len(header), MAX_PARSE_COLS)
    seen: dict[str, int] = {}
    columns = [_dedupe(header[c].strip() or f"Column {c + 1}", seen) for c in range(end_col)]
    rows: list[dict[str, Any]] = []
    for raw in collected[1:]:
        rows.append({columns[c]: _infer(raw[c]) if c < len(raw) else None for c in range(end_col)})
    total_rows = len(collected) - 1
    truncated = total_rows > len(rows) or len(header) > end_col
    return {
        "kind": "spreadsheet",
        "sheets": ["Sheet1"],
        "sheet": "Sheet1",
        "columns": columns,
        "rows": rows,
        "rowCount": len(rows),
        "totalRows": total_rows,
        "truncated": truncated,
        "truncationNote": "",
    }


def _infer(cell: str) -> Any:
    """SheetJS-style type inference for CSV (documented caveat: a leading-zero
    identifier like '007' coerces to 7 — text-formatted xlsx preserves it)."""
    if cell == "":
        return None
    try:
        return int(cell)
    except ValueError:
        pass
    try:
        return float(cell)
    except ValueError:
        return cell


def parse_word(buffer: bytes, filename: str) -> dict[str, Any]:
    # Reuse Plan A's shared docx→Markdown extraction (structural gate + mammoth).
    try:
        result = extract_office(buffer, WORD_MEDIA_TYPE, name=filename)
    except OfficeExtractError as exc:
        raise FileParseError(str(exc), status=400, code="INVALID_OFFICE_FILE") from exc
    return {
        "kind": "document",
        "format": "word",
        "text": result.text,
        "truncated": result.truncated,
        "truncationNote": result.truncation_note,
    }


def _extract_office_payload(buffer: bytes, media_type: str, filename: str) -> dict[str, Any]:
    """Chat office-extract → Markdown (the `ExtractResult` shape as a JSON-safe dict).
    Runs INSIDE the governor child so an inflate bomb is bounded by the child's rlimit."""
    try:
        result = extract_office(buffer, media_type, name=filename)
    except OfficeExtractError as exc:
        raise FileParseError(str(exc), status=400, code="INVALID_OFFICE_FILE") from exc
    return {
        "format": result.format,
        "text": result.text,
        "truncated": result.truncated,
        "truncationNote": result.truncation_note,
    }


def parse_dispatch(buffer: bytes, kind: str, filename: str, sheet: str | None) -> dict[str, Any]:
    """Run the bounds then the parser for `kind`. Called INSIDE the killable governor
    child. `__test_*` kinds are test-only governor seams (never reachable from the
    HTTP endpoint, whose `parse_kind_for` returns only real kinds). The `extract_*`
    kinds serve the chat office→Markdown path (A.U11), sharing this governor so an
    untrusted docx/xlsx inflate can never OOM the shared API worker."""
    if kind == "__test_sleep":  # governor timeout seam
        import time

        time.sleep(30)
        return {}
    if kind == "__test_crash":  # governor hard-kill seam (simulates an OOM-kill)
        import os

        os._exit(137)
    if kind == "__test_oom":  # governor memory-ceiling seam
        _ = bytearray(4 * 1024 * 1024 * 1024)
        return {}

    if kind == "extract_word":  # chat docx → Markdown, zip-bomb-bounded in the governor
        assert_zip_not_bomb(buffer)
        return _extract_office_payload(buffer, WORD_MEDIA_TYPE, filename)
    if kind == "extract_excel":  # chat xlsx → Markdown, zip-bomb-bounded in the governor
        assert_zip_not_bomb(buffer)
        return _extract_office_payload(buffer, EXCEL_MEDIA_TYPE, filename)

    if kind == "xlsx":
        try:
            assert_office_structure(buffer, "excel")
        except OfficeExtractError as exc:
            raise FileParseError(str(exc), status=400, code="INVALID_OFFICE_FILE") from exc
        assert_zip_not_bomb(buffer)
        return parse_spreadsheet(buffer, sheet)
    if kind == "word":
        assert_zip_not_bomb(buffer)
        return parse_word(buffer, filename)
    if kind == "csv":
        if looks_like_zip(buffer):
            assert_zip_not_bomb(buffer)
        return parse_csv(buffer, sheet)
    if kind == "xls":
        if looks_like_zip(buffer):
            assert_zip_not_bomb(buffer)
        return parse_spreadsheet(buffer, sheet)
    raise FileParseError(
        "Supported: Excel (.xlsx/.xls), CSV, Word (.docx).",
        status=415,
        code="UNSUPPORTED_TYPE",
    )
